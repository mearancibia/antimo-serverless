# -*- coding: utf-8 -*-
"""Capa de acceso a datos para el motor (engine.compute). Dos implementaciones que producen el
MISMO dict `src`:

- LocalSource:    lee del Excel base + rankings de entrada/ + datos/*.json. Sirve para (a) verificar
                  que el motor refactorizado da idéntico al actualizar_antimo.py original, y (b)
                  seguir corriendo el motor en la Mac. Necesita openpyxl.
- SupabaseSource: lee de las tablas de Supabase. Es la que usan las funciones serverless en /api.
                  NO necesita openpyxl (los datos maestros ya están en la DB).

El contrato de `src` está documentado en engine.py.
"""
import os, re, json, glob, base64, mimetypes, tempfile, zipfile, shutil, contextlib

from engine import norm


# ================================================================ CAJA DE RESPALDO (fusión)
# Campos de una caja que se SUMAN al fusionar dos fuentes de la misma noche, y los que se
# concatenan. Espejo de bistro._nuevo_dia(): si se agrega un campo allá, agregarlo acá.
_CAJA_SUMAR = ("total_vendido", "efectivo", "tarjetas", "qr", "otros_pago",
               "comensales", "descuentos", "retiros", "depositos")
_CAJA_CONCAT = ("detalle_retiros", "detalle_descuentos")


def _caja_key(c):
    return str(c.get("fecha_key") or c.get("fecha_iso") or c.get("fecha") or "")


def fusionar_backup(ventas, cajas, ventas_bk, cajas_bk, excluidos):
    """Suma la caja de respaldo a los datos de Bistrosoft. Devuelve (ventas, cajas) nuevas.

    Son dos fuentes INDEPENDIENTES que se suman, no un merge con resolución de conflictos:
    una venta se cobra en Bistrosoft O en la caja de respaldo, nunca en las dos (§ válvula).

    - ventas: se concatenan las filas y listo. El motor agrupa por (norm(nombre), iso) y suma,
      así que un producto que ya vendió Bistrosoft esa noche cae en la MISMA fila. Por eso el
      nombre tiene que viajar crudo desde el cliente: normalizarlo antes lo desdoblaría.
    - cajas: se suman ACÁ, antes del motor, para que engine._dedup_cajas siga recibiendo una
      sola caja por noche y no haya que tocarlo. La caja fusionada conserva el `archivo` de la
      fuente Bistrosoft, así la preferencia del dedupe frente al PDF queda intacta.

    `excluidos` (set de ISO) es la válvula: saca del cómputo ventas Y caja de esa noche. Las dos
    juntas — sacar solo la caja dejaría los productos contados doble.
    """
    ex = set(excluidos or ())
    ventas = list(ventas or []) + [v for v in (ventas_bk or []) if str(v.get("iso") or "") not in ex]

    bk = {}
    for c in (cajas_bk or []):
        k = _caja_key(c)
        if k and k not in ex:
            bk[k] = c
    if not bk:
        return ventas, list(cajas or [])

    out = []
    for c in (cajas or []):
        k = _caja_key(c)
        b = bk.pop(k, None)
        if b is None:
            out.append(c)
            continue
        m = dict(c)
        for f in _CAJA_SUMAR:
            m[f] = (c.get(f) or 0) + (b.get(f) or 0)
        for f in _CAJA_CONCAT:
            m[f] = list(c.get(f) or []) + list(b.get(f) or [])
        out.append(m)          # conserva fecha/fecha_key/archivo de la fuente Bistrosoft
    # noches que SOLO tienen caja de respaldo (el bar cobró todo por la segunda caja)
    out += list(bk.values())
    return ventas, out


# ================================================================ LOCAL (Excel + archivos)
class LocalSource:
    def __init__(self, base=None):
        self.BASE = base or os.path.dirname(os.path.abspath(__file__))
        self.DATOS = os.path.join(self.BASE, "datos")
        self.ENTRADA = os.path.join(self.BASE, "entrada")

    # -- helpers de archivos --
    def _find(self, patterns, folders):
        for fo in folders:
            for pat in patterns:
                hits = sorted(glob.glob(os.path.join(fo, pat)), key=os.path.getmtime)
                if hits:
                    return hits[-1]
        return None

    def _find_all(self, patterns, folders):
        out = []
        for fo in folders:
            for pat in patterns:
                for h in glob.glob(os.path.join(fo, pat)):
                    if h not in out:
                        out.append(h)
        return sorted(out)

    def _load_json(self, name, default):
        """Un archivo ilegible ABORTA la corrida. No devuelve el default.

        Invariante 3 del CLAUDE.md, mismo criterio que `app_antimo._load(estricto=True)`.
        Tragarse la excepción y seguir con el default tenía dos consecuencias, las dos mudas:

        - Con `opex.json` corrupto, el motor lo veía como "nunca sembrado" y el driver escribía
          la siembra ENCIMA: el archivo dañado (recuperable desde datos/_backups/) se perdía
          para siempre y el tablero pasaba a mostrar el OPEX de julio como si fuera el de hoy.
        - Con cualquier otro override corrupto, sus datos desaparecían del tablero sin aviso:
          los márgenes seguían mostrando números creíbles, calculados sin esos overrides.

        Parar y decir qué archivo es siempre mejor que seguir con números equivocados.
        """
        p = os.path.join(self.DATOS, name)
        if not os.path.exists(p):
            return default
        try:
            with open(p, encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            raise SystemExit(
                "ERROR: datos/%s está corrupto (%s).\n"
                "No se recalculó nada, para no pisarlo con datos incompletos.\n"
                "Hay copias de las últimas 20 versiones en datos/_backups/ — restaurá la más "
                "reciente que abra bien y volvé a correr." % (name, e))

    @contextlib.contextmanager
    def _ranking_reparado(self, src):
        tmp = tempfile.mkdtemp(prefix="antimo_rank_")
        try:
            dst = os.path.join(tmp, "rank.xlsx")
            with zipfile.ZipFile(src) as z:
                z.extractall(tmp)
            ws = os.path.join(tmp, "xl", "worksheets", "sheet1.xml")
            if os.path.exists(ws):
                with open(ws, encoding="utf-8") as f:
                    x = f.read()
                x = re.sub(r"<mergeCells.*?</mergeCells>", "", x, flags=re.S)
                x = re.sub(r"<mergeCell [^>]*/>", "", x)
                with open(ws, "w", encoding="utf-8") as f:
                    f.write(x)
            with zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as z:
                for root, _, files in os.walk(tmp):
                    for f in files:
                        fp = os.path.join(root, f)
                        if fp != dst:
                            z.write(fp, os.path.relpath(fp, tmp))
            yield dst
        finally:
            shutil.rmtree(tmp, ignore_errors=True)

    def build(self):
        import openpyxl
        DG = self._find(["datos_general.xlsx", "datos general.xlsx", "*general*.xlsx"], [self.DATOS, self.BASE])
        if not DG:
            raise SystemExit("Falta datos/datos_general.xlsx")
        wb = openpyxl.load_workbook(DG, data_only=True)

        # ---- Costo_Base ----
        costo_base = {}; cb_cat = {}
        for r in list(wb["Costo_Base"].iter_rows(values_only=True))[1:]:
            if r[1] is None:
                continue
            costo_base[r[1]] = {"precio": r[2], "pres": r[3], "cant_base": r[4], "unidad": r[5], "cxu": r[6]}
            cb_cat[r[1]] = r[0]

        # ---- Lista de Precios (normalizada) ----
        precio_lista = {}
        if "Lista de Precios" in wb.sheetnames:
            for r in list(wb["Lista de Precios"].iter_rows(values_only=True))[1:]:
                if r[1] is None or r[2] is None:
                    continue
                precio_lista[norm(r[1])] = r[2]

        # ---- Recetas (Bebidas + Comida) ----
        recetas = {}
        for sheet, i0, nc in [("Recetas Bebidas", 1, 0), ("Recetas Comida", 2, 1)]:
            for r in list(wb[sheet].iter_rows(values_only=True))[1:]:
                cells = list(r); nm = cells[nc]
                if not nm:
                    continue
                ings = []
                for j in range(i0, len(cells) - 1, 2):
                    ing, qty = cells[j], cells[j + 1]
                    if ing and str(ing).strip():
                        ings.append((str(ing).strip(), qty))
                recetas[norm(nm)] = ings

        # ---- Maestro_Productos ----
        maestro = {}
        for r in list(wb["Maestro_Productos"].iter_rows(values_only=True))[1:]:
            if r[0] is None:
                continue
            maestro[norm(r[0])] = {"cat": r[1], "canon": r[2], "tipo": r[3],
                                   "factor": r[4] or 1, "rend": r[5], "costeo": r[6], "nota": r[7]}

        # ---- OPEX base (para sembrar si opex.json no existe) ----
        opex_base = []
        if "Costos FIJOS (OPEX)" in wb.sheetnames:
            for r in wb["Costos FIJOS (OPEX)"].iter_rows(min_row=2, values_only=True):
                if not r or r[0] is None or r[1] is None:
                    continue
                opex_base.append({"cat": r[0], "item": r[1], "cantidad": r[2], "unitario": r[3],
                                  "monto": r[5] if len(r) > 5 else None})

        # ---- ventas (rankings de entrada/, año desde el nombre de archivo) ----
        RANKS = self._find_all(["*anking*.xlsx", "*ent*.xlsx", "*Vent*.xlsx", "api_*.xlsx"], [self.ENTRADA])
        ventas = []
        for path in RANKS:
            m = re.search(r"(20\d\d)-(\d\d)", os.path.basename(path))
            ym = (m.group(1), m.group(2)) if m else None
            try:
                with self._ranking_reparado(path) as _rp:
                    rows = list(openpyxl.load_workbook(_rp, data_only=True).active.iter_rows(values_only=True))
            except Exception:
                continue
            hdr = None; header = []
            for i, row in enumerate(rows[:6]):
                vals = [str(c).strip().lower() if c is not None else "" for c in row]
                if "nombre" in vals and ("vendidos" in vals or "monto" in vals):
                    hdr = i; header = vals; break
            if hdr is None:
                continue

            def col(*n):
                for x in n:
                    if x in header:
                        return header.index(x)
                return None
            cN, cV, cM, cF = col("nombre"), col("vendidos"), col("monto"), col("fecha")
            for r in rows[hdr + 1:]:
                if not r or cN is None or r[cN] is None:
                    continue
                fecha = str(r[cF]) if cF is not None and r[cF] is not None else "s/f"
                dd, mm = (fecha.split("-") + ["", ""])[:2]
                yy = ym[0] if ym else ""
                iso = f"{yy}-{mm}-{dd}" if (yy and mm and dd) else ""
                ventas.append({"nombre": r[cN], "fecha": fecha, "iso": iso,
                               "unidades": r[cV] if cV is not None else 0,
                               "monto": r[cM] if cM is not None else 0})

        # ---- cajas (API json) ----
        cajas = self._load_json("cajas_api.json", [])

        # ---- caja de respaldo (opcional: en local casi siempre no existe) ----
        # Se leen igual que en la nube para que el gate de verificación (compute(LocalSource) ==
        # compute(SupabaseSource)) siga comparando lo mismo. Sin archivos → listas vacías → la
        # fusión es un no-op y el DATA queda idéntico a antes.
        ventas, cajas = fusionar_backup(
            ventas, cajas,
            self._load_json("ventas_backup.json", []),
            self._load_json("cajas_backup.json", []),
            self._load_json("backup_excluido.json", []))

        # ---- OPEX json (editable) ----
        opex_json = self._load_json("opex.json", None)

        # ---- logo ----
        logo = ""
        for n in ["logo.png", "logo.jpg", "logo.jpeg", "logo.svg", "logo.webp"]:
            for fo in [self.BASE, self.DATOS]:
                lp = os.path.join(fo, n)
                if os.path.exists(lp):
                    mt = mimetypes.guess_type(lp)[0] or "image/png"
                    logo = "data:%s;base64,%s" % (mt, base64.b64encode(open(lp, "rb").read()).decode())
                    break
            if logo:
                break

        return {
            "costo_base": costo_base, "cb_cat": cb_cat, "precio_lista": precio_lista,
            "recetas": recetas, "maestro": maestro, "ventas": ventas, "cajas": cajas,
            "opex_json": opex_json, "opex_base": opex_base,
            "opex_cero_confirmado": self._load_json("opex_cero_confirmado.json", []),
            "overrides": {
                "maestro_extra": self._load_json("maestro_extra.json", []),
                "pours_extra": self._load_json("pours_extra.json", {}),
                "recetas_extra": self._load_json("recetas_extra.json", {}),
                "insumos_extra": self._load_json("insumos_extra.json", []),
                "precios_override": self._load_json("precios_override.json", {}),
                "combos_extra": self._load_json("combos_extra.json", {}),
                "precio_lista_override": self._load_json("precio_lista_override.json", {}),
                "sospechosos": self._load_json("sospechosos.json", {}),
                "dias_cerrados": self._load_json("dias_cerrados.json", {}),
                "stock": self._load_json("stock.json", {}),
            },
            "logo": logo,
        }


# ================================================================ SUPABASE
class SupabaseSource:
    """Lee todo de Supabase. Las tablas de datos maestros (costo_base, lista_precios, recetas,
    maestro_productos, opex_base, ventas, cajas) se siembran una vez con scripts/seed_supabase.py.
    Las tablas de override las escriben los endpoints de /api. `app_meta` guarda el logo (data-URI)
    y el opex.json editable como filas de una tabla clave-valor JSON."""

    def __init__(self, client):
        self.sb = client

    def _rows(self, table, cols="*"):
        # PostgREST corta las SELECT en 1000 filas por defecto: sin paginar, la tabla `ventas`
        # (>1000 filas) se leía truncada y se perdían los meses más nuevos EN SILENCIO. Se pagina
        # con .range() hasta agotar. Sin ordenar explícito, PostgREST puede repetir/saltear filas
        # entre páginas, así que ordenamos por la PK de cada tabla (o por una columna estable).
        order_col = {"ventas": "id", "opex_base": "id", "ventas_backup": "id"}.get(table)
        out = []; start = 0; PAGE = 1000
        while True:
            q = self.sb.table(table).select(cols)
            if order_col:
                q = q.order(order_col)
            batch = q.range(start, start + PAGE - 1).execute().data or []
            out += batch
            if len(batch) < PAGE:
                break
            start += PAGE
        return out

    def _kv(self, table, key_col, val_col):
        return {r[key_col]: r[val_col] for r in self._rows(table)}

    def _meta(self, key, default=None):
        r = self.sb.table("app_meta").select("value").eq("key", key).execute().data
        return r[0]["value"] if r else default

    def build(self):
        costo_base = {}; cb_cat = {}
        for r in self._rows("costo_base"):
            costo_base[r["nombre"]] = {"precio": r.get("precio"), "pres": r.get("pres"),
                "cant_base": r.get("cant_base"), "unidad": r.get("unidad"), "cxu": r.get("cxu")}
            cb_cat[r["nombre"]] = r.get("cb_cat")

        precio_lista = {}
        for r in self._rows("lista_precios"):
            precio_lista[norm(r["nombre"])] = r.get("precio")

        recetas = {}
        for r in self._rows("recetas"):
            recetas[norm(r["nombre"])] = [(i[0], i[1]) for i in (r.get("ingredientes") or [])]

        maestro = {}
        for r in self._rows("maestro_productos"):
            maestro[norm(r["pos"])] = {"cat": r.get("cat"), "canon": r.get("canon"),
                "tipo": r.get("tipo"), "factor": r.get("factor") or 1, "rend": r.get("rend"),
                "costeo": r.get("costeo"), "nota": r.get("nota")}

        opex_base = [{"cat": r.get("cat"), "item": r.get("item"), "cantidad": r.get("cantidad"),
                      "unitario": r.get("unitario"), "monto": r.get("monto")}
                     for r in self._rows("opex_base")]

        ventas = [{"nombre": r.get("nombre"), "fecha": r.get("fecha"), "iso": r.get("iso"),
                   "unidades": r.get("unidades"), "monto": r.get("monto")}
                  for r in self._rows("ventas")]

        cajas = [r["data"] for r in self._rows("cajas")]

        # ---- caja de respaldo: se SUMA a lo de Bistrosoft (ver fusionar_backup) ----
        ventas_bk = [{"nombre": r.get("nombre"), "fecha": r.get("fecha"), "iso": r.get("iso"),
                      "unidades": r.get("unidades"), "monto": r.get("monto")}
                     for r in self._rows("ventas_backup")]
        cajas_bk = [r["data"] for r in self._rows("cajas_backup")]
        excluidos = [r["iso"] for r in self._rows("backup_excluido")]
        ventas, cajas = fusionar_backup(ventas, cajas, ventas_bk, cajas_bk, excluidos)

        # override tables (dicts/lists)
        maestro_extra = [r["data"] for r in self._rows("maestro_extra")]
        insumos_extra = [r["data"] for r in self._rows("insumos_extra")]
        recetas_extra = {r["nombre"]: r["ingredientes"] for r in self._rows("recetas_extra")}
        combos_extra = {r["pos"]: r["componentes"] for r in self._rows("combos_extra")}
        pours_extra = self._kv("pours_extra", "key", "rend")
        precios_override = self._kv("precios_override", "insumo", "precio")
        precio_lista_override = self._kv("precio_lista_override", "key", "precio")
        stock = {r["insumo"]: r["data"] for r in self._rows("stock")}
        sospechosos = {r["key"]: r["data"] for r in self._rows("sospechosos")}
        dias_cerrados = self._kv("dias_cerrados", "iso", "motivo")

        return {
            "costo_base": costo_base, "cb_cat": cb_cat, "precio_lista": precio_lista,
            "recetas": recetas, "maestro": maestro, "ventas": ventas, "cajas": cajas,
            "opex_json": self._meta("opex_json", None), "opex_base": opex_base,
            "opex_cero_confirmado": self._meta("opex_cero_confirmado", []),
            "overrides": {
                "maestro_extra": maestro_extra, "pours_extra": pours_extra,
                "recetas_extra": recetas_extra, "insumos_extra": insumos_extra,
                "precios_override": precios_override, "combos_extra": combos_extra,
                "precio_lista_override": precio_lista_override, "sospechosos": sospechosos,
                "dias_cerrados": dias_cerrados, "stock": stock,
            },
            "logo": self._meta("logo", "") or "",
        }
