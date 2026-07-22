#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""ANTIMO - App local. Sirve el tablero y permite editar OPEX, recetas, costos y combos,
guardando en archivos override y recalculando al instante. Sin dependencias externas."""
import http.server, socketserver, json, os, sys, subprocess, threading, webbrowser, urllib.parse
import glob, shutil, datetime, re
BASE=os.path.dirname(os.path.abspath(__file__)); DATOS=os.path.join(BASE,"datos")
BACKUPS=os.path.join(DATOS,"_backups")
BACKUP_KEEP=20          # versiones que se conservan por archivo
NO_BACKUP={"bistro_config.json"}   # credenciales: no multiplicar copias en claro
PORT=8733
MAX_BODY=8*1024*1024    # tope del cuerpo de un POST (el import de stock mas grande no llega a 1MB)
# Los handlers hacen read-modify-write sobre los JSON y el server es multi-hilo: sin esto, dos
# guardados solapados (tabular rapido entre inputs de OPEX ya los genera) pierden una edicion.
_LOCK=threading.RLock()
class DatosCorruptos(Exception):
    """Un override existe pero no se puede parsear. NO es lo mismo que 'todavia no existe':
    devolver el default en ese caso hace que el siguiente _save lo pise con un solo item."""
    pass
def _load(name,default,estricto=False):
    """estricto=True para todo read-modify-write: ante un archivo ilegible aborta el guardado
    en vez de arrancar de cero y borrar lo que habia."""
    p=os.path.join(DATOS,name)
    if not os.path.exists(p): return default
    try:
        with open(p,encoding="utf-8") as f: return json.load(f)
    except Exception as e:
        if estricto:
            raise DatosCorruptos(f"{name} está corrupto ({e}). No se guardó nada para no "
                                 f"pisar los datos. Hay copias en datos/_backups/.")
        print("WARN: no pude leer",name,"->",e)
        return default
def _backup(name):
    """Copia la version previa a datos/_backups/ antes de pisarla. Nunca interrumpe el guardado."""
    if name in NO_BACKUP: return
    p=os.path.join(DATOS,name)
    if not os.path.exists(p): return
    try:
        os.makedirs(BACKUPS,exist_ok=True)
        stem,ext=os.path.splitext(name)
        stamp=datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
        shutil.copy2(p,os.path.join(BACKUPS,stem+"."+stamp+ext))
        viejos=sorted(glob.glob(os.path.join(BACKUPS,glob.escape(stem)+".*"+ext)))
        for f in viejos[:-BACKUP_KEEP]:
            try: os.remove(f)
            except OSError: pass
    except Exception as e:
        print("WARN: no pude respaldar",name,"->",e)
def _save(name,obj):
    """Escritura ATOMICA: tmp + os.replace. open(...,'w') trunca el archivo antes de escribir,
    asi que si el proceso muere en el medio (cerrar la ventana es la forma documentada de cerrar
    la app) queda un JSON cortado — que es justo lo que _load no sabe distinguir de uno nuevo."""
    with _LOCK:
        _backup(name)
        p=os.path.join(DATOS,name); tmp=p+".tmp"
        with open(tmp,"w",encoding="utf-8") as f:
            json.dump(obj,f,ensure_ascii=False,indent=1)
            f.flush(); os.fsync(f.fileno())
        os.replace(tmp,p)
OPEX_DESDE_0="2000-01-01"
def _opex_periodos():
    """opex.json normalizado a vigencias [{desde,items}], ordenadas. El formato viejo
    (lista plana de rubros) se lee como una unica vigencia que rige desde siempre."""
    raw=_load("opex.json",[],estricto=True)
    if not raw: return [{"desde":OPEX_DESDE_0,"items":[]}]
    if isinstance(raw[0],dict) and "items" in raw[0]:
        return sorted([p for p in raw if isinstance(p,dict)],key=lambda p:str(p.get("desde") or OPEX_DESDE_0))
    return [{"desde":OPEX_DESDE_0,"items":raw}]
def _opex_vigente(ps):
    hoy=datetime.date.today().isoformat()
    cur=[p for p in ps if str(p.get("desde") or OPEX_DESDE_0)<=hoy]
    return str((cur[-1] if cur else ps[0]).get("desde") or OPEX_DESDE_0)
def run_pipeline():
    r=subprocess.run([sys.executable,os.path.join(BASE,"actualizar_antimo.py")],cwd=BASE,capture_output=True,text=True)
    return r.returncode==0, (r.stdout or "")+(r.stderr or "")
def run_connector():
    r=subprocess.run([sys.executable,os.path.join(BASE,"conector_bistrosoft.py")],cwd=BASE,capture_output=True,text=True)
    return r.returncode==0, (r.stdout or "")+(r.stderr or "")


def generate_excel():
    import openpyxl, shutil, datetime
    src=os.path.join(DATOS,"datos_general.xlsx"); dst=os.path.join(BASE,"datos_general_actualizado.xlsx")
    if not os.path.exists(src): return None
    wb=openpyxl.load_workbook(src)
    pov=_load("precios_override.json",{})
    if pov and "Costo items" in wb.sheetnames:
        for row in wb["Costo items"].iter_rows(min_row=2):
            if row[1].value in pov: row[2].value=pov[row[1].value]
    opx=_load("opex.json",[])
    if opx and "Costos FIJOS (OPEX)" in wb.sheetnames:
        ws=wb["Costos FIJOS (OPEX)"]; existing={}
        for row in ws.iter_rows(min_row=2):
            if row[1].value: existing[str(row[1].value).strip().lower()]=row
        for e in opx:
            key=str(e.get("item","")).strip().lower(); cant=e.get("cantidad") or 0; unit=e.get("unitario") or 0; monto=round(cant*unit)
            if key in existing:
                r=existing[key]; r[2].value=cant; r[3].value=unit; r[5].value=monto
            else:
                ws.append([e.get("cat"),e.get("item"),cant,unit,"(agregado en app)",monto])
    # hojas extra para no perder nada
    def _sheet(name):
        if name in wb.sheetnames: del wb[name]
        return wb.create_sheet(name)
    rex=_load("recetas_extra.json",{})
    if rex:
        def _n(x): return ' '.join(str(x).strip().upper().split()) if x else ''
        rem=dict(rex)
        for sheet,namecol,ingstart in [("Recetas Bebidas",0,1),("Recetas Comida",1,2)]:
            if sheet not in wb.sheetnames: continue
            ws=wb[sheet]
            for row in ws.iter_rows(min_row=2):
                nm=_n(row[namecol].value)
                for rk in list(rem.keys()):
                    if _n(rk)==nm:
                        ings=rem.pop(rk); r=row[0].row; ci=ingstart
                        for ig in ings:
                            ws.cell(row=r,column=ci+1,value=ig[0]); ws.cell(row=r,column=ci+2,value=ig[1]); ci+=2
                        for c in range(ci+1,ws.max_column+1): ws.cell(row=r,column=c,value=None)
                        break
        if rem:  # recetas nuevas que no estaban en las hojas
            ws=_sheet("Recetas_nuevas"); ws.append(["Receta","Ingrediente","Cantidad"])
            for nm,ings in rem.items():
                for ig in ings: ws.append([nm,ig[0],ig[1]])
    mex=_load("maestro_extra.json",[])
    if mex:
        ws=_sheet("Productos_agregados"); ws.append(["POS","Categoria","Canonico","Tipo","Factor","Rend","Costeo","Nota"])
        for e in mex: ws.append([e.get("pos"),e.get("cat"),e.get("canon"),e.get("tipo"),e.get("factor"),e.get("rend"),e.get("costeo"),e.get("nota")])
    iex=_load("insumos_extra.json",[])
    if iex:
        ws=_sheet("Insumos_agregados"); ws.append(["Nombre","Precio","Presentacion","Cant_base","Unidad","cxu"])
        for e in iex: ws.append([e.get("nombre"),e.get("precio"),e.get("pres"),e.get("cant_base"),e.get("unidad"),e.get("cxu")])
    wb.save(dst); return dst

# Pantalla de primer arranque (instalacion nueva, sin ventas todavia). Mismo criterio que el
# resto: un solo archivo, sin dependencias, sin nada externo.
PAGINA_PRIMER_ARRANQUE="""<!DOCTYPE html><html lang="es"><head><meta charset="utf-8">
<title>ANTIMO · Primeros pasos</title><style>
body{background:#0f1419;color:#e8edf2;font:15px/1.6 -apple-system,Segoe UI,Roboto,sans-serif;
 max-width:620px;margin:0 auto;padding:50px 24px}
h1{font-size:26px;margin-bottom:6px}
.sub{color:#8a97a6;margin-bottom:28px}
.card{background:#1a222c;border:1px solid #2b3644;border-radius:12px;padding:20px 22px;margin-bottom:14px}
.paso{display:flex;gap:14px;align-items:flex-start}
.num{background:#e8834a;color:#151515;font-weight:700;border-radius:50%;width:26px;height:26px;
 display:flex;align-items:center;justify-content:center;flex:0 0 26px;font-size:13px}
label{display:block;font-size:11px;color:#8a97a6;text-transform:uppercase;letter-spacing:.4px;margin:10px 0 3px}
input{background:#212b37;color:#e8edf2;border:1px solid #2b3644;border-radius:8px;padding:9px 11px;
 font-size:14px;width:100%}
button{background:#e8834a;color:#151515;border:0;border-radius:8px;padding:11px 18px;font-size:14px;
 font-weight:600;cursor:pointer;margin-top:16px}
button:disabled{opacity:.5;cursor:default}
#msg{margin-top:14px;font-size:13.5px;min-height:20px}
.ok{color:#4a9d7f}.err{color:#c85a54}
</style></head><body>
<h1>Bienvenido a ANTIMO</h1>
<div class="sub">Falta un paso para empezar: conectar tu cuenta de Bistrosoft y traer las ventas.</div>
<div class="card"><div class="paso"><div class="num">1</div><div style="flex:1">
 <b>Conecta tu cuenta</b>
 <div style="color:#8a97a6;font-size:13.5px">Son los mismos datos con los que entras a Bistrosoft.
  Quedan guardados solo en esta computadora.</div>
 <label>Usuario (email)</label><input id="u" autocomplete="off">
 <label>Contrase&ntilde;a</label><input id="p" type="password" autocomplete="off">
 <label>C&oacute;digo de tienda (shopCode)</label><input id="s" autocomplete="off">
</div></div></div>
<div class="card"><div class="paso"><div class="num">2</div><div style="flex:1">
 <b>Tra&eacute; las ventas</b>
 <div style="color:#8a97a6;font-size:13.5px">La primera vez puede tardar un rato: baja todo el
  historial disponible. Despu&eacute;s el tablero abre solo.</div>
 <button id="go">Conectar y traer ventas</button>
 <div id="msg"></div>
</div></div></div>
<script>
const $=i=>document.getElementById(i);
$('go').onclick=async()=>{
 const u=$('u').value.trim(),p=$('p').value,s=$('s').value.trim();
 if(!u||!p||!s){$('msg').className='err';$('msg').textContent='Completa los tres campos.';return;}
 $('go').disabled=true;$('msg').className='';$('msg').textContent='Guardando la cuenta...';
 const post=(url,body)=>fetch(url,{method:'POST',headers:{'Content-Type':'application/json'},
   body:JSON.stringify(body)}).then(r=>r.json());
 try{
  const c=await post('/api/config',{base:'https://ar-api.bistrosoft.com',username:u,password:p,shopCode:s});
  if(!c.ok)throw new Error('No pude guardar la cuenta');
  $('msg').textContent='Trayendo las ventas... (puede tardar varios minutos, no cierres esta ventana)';
  const j=await post('/api/pull',{});
  if(j.ok){$('msg').className='ok';$('msg').textContent='Listo. Abriendo el tablero...';
   setTimeout(()=>location.reload(),1200);}
  else{$('msg').className='err';
   $('msg').textContent='No se pudieron traer las ventas: '+((j.log||j.error||'').slice(-260)||'error desconocido');
   $('go').disabled=false;}
 }catch(e){$('msg').className='err';$('msg').textContent='Error: '+e.message;$('go').disabled=false;}
};
</script></body></html>"""

class H(http.server.SimpleHTTPRequestHandler):
    def _send(self,code,body,ctype="application/json"):
        b=body.encode("utf-8") if isinstance(body,str) else body
        self.send_response(code); self.send_header("Content-Type",ctype+"; charset=utf-8")
        self.send_header("Content-Length",str(len(b))); self.send_header("Cache-Control","no-store"); self.end_headers()
        self.wfile.write(b)
    def _origen_confiable(self):
        """Un POST solo puede venir de la propia pagina de ANTIMO. Tres chequeos:
        1) Content-Type application/json: obliga al navegador a hacer preflight en cualquier
           pedido cross-origin, y el preflight falla porque no mandamos cabeceras CORS. Esto es
           lo que corta el CSRF: sin el chequeo, una web cualquiera manda text/plain (que no
           dispara preflight), el server igual hace json.loads del body, y escribe en disco.
        2) Origin: los navegadores lo mandan siempre en POST. Si viene, tiene que ser loopback.
           curl y los scripts locales no lo mandan y siguen funcionando.
        3) Host: bloquea DNS rebinding (un dominio que resuelve a 127.0.0.1 alcanza el server
           aunque este bindeado a loopback)."""
        ct=(self.headers.get("Content-Type") or "").split(";")[0].strip().lower()
        if ct!="application/json": return False
        org=self.headers.get("Origin")
        if org is not None and urllib.parse.urlparse(org).hostname not in ("127.0.0.1","localhost"):
            return False
        return (self.headers.get("Host") or "").split(":")[0] in ("127.0.0.1","localhost")
    def do_HEAD(self):
        """NO heredar el de SimpleHTTPRequestHandler: sirve toda la carpeta del proyecto y
        deja ver que existe (y cuanto pesa) datos/bistro_config.json, que do_GET bloquea."""
        self.do_GET()
    def do_GET(self):
        path=urllib.parse.urlparse(self.path).path
        if path in ("/","/index.html"):
            f=os.path.join(BASE,"dashboard_ANTIMO.html")
            if os.path.exists(f):
                with open(f,encoding="utf-8") as fh: return self._send(200,fh.read(),"text/html")
            # Instalacion nueva: todavia no hay ventas, asi que el tablero no existe. Antes esto
            # devolvia "no dashboard" y el usuario quedaba sin ningun lugar donde configurar la
            # cuenta ni traer las ventas. Se sirve una pantalla de primer arranque.
            return self._send(200,PAGINA_PRIMER_ARRANQUE,"text/html")
        if path=="/api/data":
            f=os.path.join(BASE,"datos_dashboard.json")
            return self._send(200,open(f,encoding="utf-8").read()) if os.path.exists(f) else self._send(404,'{}')
        if path=="/api/config":
            c=_load("bistro_config.json",{})
            return self._send(200,json.dumps({"base":c.get("base","https://ar-api.bistrosoft.com"),"username":c.get("username",""),"shopCode":c.get("shopCode",""),"configured":bool(c.get("username") and not str(c.get("username","")).startswith("TU_"))}))
        if path=="/api/ping": return self._send(200,'{"app":true}')
        return self._send(404,"not found")
    def do_POST(self):
        if not self._origen_confiable():
            return self._send(403,'{"ok":false,"error":"origen no permitido"}')
        path=urllib.parse.urlparse(self.path).path
        # el int() estaba FUERA del try: un Content-Length no numerico mataba el hilo sin
        # responder nada, y el navegador quedaba esperando hasta el timeout.
        try:
            ln=int(self.headers.get("Content-Length") or 0)
            if ln<0 or ln>MAX_BODY: raise ValueError("tamaño fuera de rango")
        except ValueError:
            return self._send(400,'{"ok":false,"error":"Content-Length inválido"}')
        try: data=json.loads(self.rfile.read(ln) or b"{}")
        except Exception: data={}
        # Serializa las mutaciones. El server es multi-hilo y todos los handlers hacen
        # read-modify-write sobre los mismos JSON: dos POST solapados pierden una edicion.
        with _LOCK:
            return self._post_dispatch(path,data)
    def _post_dispatch(self,path,data):
        try:
            if path=="/api/config":
                c=_load("bistro_config.json",{},estricto=True)
                c["base"]=data.get("base") or c.get("base") or "https://ar-api.bistrosoft.com"
                c["username"]=data.get("username","").strip() or c.get("username","")
                if data.get("password"): c["password"]=data["password"]
                c["shopCode"]=str(data.get("shopCode","")).strip() or c.get("shopCode","")
                _save("bistro_config.json",c)
                return self._send(200,json.dumps({"ok":True}))
            if path=="/api/receta":
                r=_load("recetas_extra.json",{},estricto=True); r[data["receta"]]=[[i[0],i[1]] for i in data["ingredientes"]]; _save("recetas_extra.json",r)
            elif path=="/api/precio":
                p=_load("precios_override.json",{},estricto=True); p[data["insumo"]]=float(data["precio"]); _save("precios_override.json",p)
            elif path=="/api/opex":
                o=_load("opex_override.json",{},estricto=True); o[data["item"]]=float(data["monto"]); _save("opex_override.json",o)
            elif path=="/api/combo":
                comp=data.get("componentes") or []
                if not comp: return self._send(200,json.dumps({"ok":False,"error":"El combo quedó sin componentes"}))
                c=_load("combos_extra.json",{},estricto=True); c[data["producto"]]=[[x[0],x[1],x[2]] for x in comp]; _save("combos_extra.json",c)
            elif path=="/api/pour":
                # rendimiento (ml) de un pour. rend vacio/0 => vuelve al valor del Excel.
                key=str(data.get("key","")).strip()
                if not key: return self._send(200,json.dumps({"ok":False,"error":"Falta el producto"}))
                rend=data.get("rend")
                p=_load("pours_extra.json",{},estricto=True)
                if rend in (None,"",0,"0"): p.pop(key,None)
                else:
                    try: rend=float(rend)
                    except (TypeError,ValueError): return self._send(200,json.dumps({"ok":False,"error":"Rendimiento inválido"}))
                    if rend<=0: return self._send(200,json.dumps({"ok":False,"error":"Rendimiento inválido"}))
                    p[key]=rend
                _save("pours_extra.json",p)
            elif path=="/api/dia_cerrado":
                # el dueño marca/desmarca una noche sin apertura (asi deja de contar como dato faltante)
                iso=str(data.get("iso","")).strip()
                if not re.match(r"^\d{4}-\d{2}-\d{2}$",iso):
                    return self._send(200,json.dumps({"ok":False,"error":"Fecha inválida"}))
                c=_load("dias_cerrados.json",{},estricto=True)
                if data.get("cerrado"): c[iso]=data.get("motivo","") or "Cerrado"
                else: c.pop(iso,None)
                _save("dias_cerrados.json",c)
            elif path=="/api/stock":
                # conteo manual de un insumo vigilado. cantidad vacia/0 => dejar de vigilarlo.
                insumo=str(data.get("insumo","")).strip()
                if not insumo: return self._send(200,json.dumps({"ok":False,"error":"Falta el insumo"}))
                s=_load("stock.json",{},estricto=True)
                cant=data.get("cantidad")
                if cant in (None,"",0,"0"): s.pop(insumo,None)
                else:
                    try: cant=float(cant)
                    except (TypeError,ValueError): return self._send(200,json.dumps({"ok":False,"error":"Cantidad inválida"}))
                    if cant<0: return self._send(200,json.dumps({"ok":False,"error":"Cantidad inválida"}))
                    fecha=str(data.get("fecha") or "").strip()
                    if not re.match(r"^\d{4}-\d{2}-\d{2}$",fecha): fecha=datetime.date.today().isoformat()
                    entry={"cant":cant,"fecha":fecha}
                    umb=data.get("umbral_dias")
                    if umb not in (None,"",0,"0"):
                        try: entry["umbral_dias"]=float(umb)
                        except (TypeError,ValueError): pass
                    s[insumo]=entry
                _save("stock.json",s)
            elif path=="/api/stock_bulk":
                # importacion masiva desde el CSV exportado. Una sola fecha para todo el lote
                # (el conteo fisico se hace un dia puntual). Filas invalidas se saltean, no
                # aborta todo el import por una fila mal formada.
                fecha=str(data.get("fecha") or "").strip()
                if not re.match(r"^\d{4}-\d{2}-\d{2}$",fecha): fecha=datetime.date.today().isoformat()
                s=_load("stock.json",{},estricto=True); aplicadas=0
                for row in (data.get("items") or []):
                    insumo=str(row.get("insumo","")).strip()
                    cant=row.get("cantidad")
                    if not insumo or cant in (None,"",0,"0"): continue
                    try: cant=float(cant)
                    except (TypeError,ValueError): continue
                    if cant<=0: continue
                    entry={"cant":cant,"fecha":fecha}
                    umb=row.get("umbral_dias")
                    if umb not in (None,"",0,"0"):
                        try: entry["umbral_dias"]=float(umb)
                        except (TypeError,ValueError): pass
                    s[insumo]=entry; aplicadas+=1
                _save("stock.json",s)
            elif path=="/api/costos_bulk":
                # Importacion masiva desde el CSV de Costos: precios y/o conteos de stock, en UNA
                # sola pasada. De a uno serian ~370ms por item (cada POST re-corre el motor): con
                # 120 insumos son 43 segundos con la app aparentemente colgada.
                # El frontend ya mostro la vista previa y el dueño confirmo; aca se revalida todo
                # igual, porque el endpoint es alcanzable sin pasar por esa pantalla.
                pv=data.get("precios") or []; sk=data.get("stock") or []
                if not pv and not sk:
                    return self._send(400,json.dumps({"ok":False,"error":"No vino ningún cambio"}))
                nprec=0; nstk=0
                if pv:
                    p=_load("precios_override.json",{},estricto=True)
                    for e in pv:
                        nm=str(e.get("insumo","")).strip()
                        try: val=float(e.get("precio"))
                        except (TypeError,ValueError): continue
                        # 0 o negativo NO se acepta: a diferencia del stock, aca no significa
                        # "dejar de vigilar" — un insumo gratis falsea todos los margenes.
                        if not nm or val<=0: continue
                        p[nm]=val; nprec+=1
                    _save("precios_override.json",p)
                if sk:
                    fecha=str(data.get("fecha") or "").strip()
                    if not re.match(r"^\d{4}-\d{2}-\d{2}$",fecha): fecha=datetime.date.today().isoformat()
                    s=_load("stock.json",{},estricto=True)
                    for e in sk:
                        nm=str(e.get("insumo","")).strip()
                        try: cant=float(e.get("cantidad"))
                        except (TypeError,ValueError): continue
                        if not nm or cant<=0: continue
                        entry={"cant":cant,"fecha":fecha}
                        try:
                            u=float(e.get("umbral_dias"))
                            if u>0: entry["umbral_dias"]=u
                        except (TypeError,ValueError): pass
                        s[nm]=entry; nstk+=1
                    _save("stock.json",s)
                print(f"Importacion masiva: {nprec} precios, {nstk} conteos de stock")
            elif path=="/api/precio_lista":
                # editar el precio de lista de un producto sin tocar el Excel. vacio/0 => volver
                # al valor de la hoja "Lista de Precios" (si la tenia).
                key=str(data.get("key","")).strip()
                if not key: return self._send(200,json.dumps({"ok":False,"error":"Falta el producto"}))
                p=_load("precio_lista_override.json",{},estricto=True)
                precio=data.get("precio")
                if precio in (None,"",0,"0"): p.pop(key,None)
                else:
                    try: precio=float(precio)
                    except (TypeError,ValueError): return self._send(200,json.dumps({"ok":False,"error":"Precio inválido"}))
                    if precio<0: return self._send(200,json.dumps({"ok":False,"error":"Precio inválido"}))
                    p[key]=precio
                _save("precio_lista_override.json",p)
            elif path=="/api/sospechoso":
                # marca manual de precio mal cargado en el POS. estado: "si" | "no" | "" (limpiar)
                key=str(data.get("key","")).strip()
                if not key: return self._send(200,json.dumps({"ok":False,"error":"Falta el producto"}))
                s=_load("sospechosos.json",{},estricto=True); est=str(data.get("estado","") or "")
                if est not in ("si","no"): s.pop(key,None)
                else: s[key]={"estado":est,"motivo":data.get("motivo","") or "",
                              "ts":datetime.datetime.now().isoformat(timespec="seconds")}
                _save("sospechosos.json",s)
            elif path=="/api/opex_save":
                # guarda los rubros de UNA vigencia (sin 'desde' => la que rige hoy)
                ps=_opex_periodos()
                desde=str(data.get("desde") or "").strip() or _opex_vigente(ps)
                hit=[p for p in ps if str(p.get("desde"))==desde]
                if hit: hit[0]["items"]=data.get("items",[])
                else:
                    ps.append({"desde":desde,"items":data.get("items",[])})
                    ps.sort(key=lambda p:str(p.get("desde")))
                _save("opex.json",ps)
            elif path=="/api/opex_vigencia":
                desde=str(data.get("desde") or "").strip()
                if not re.match(r"^\d{4}-\d{2}-\d{2}$",desde):
                    return self._send(200,json.dumps({"ok":False,"error":"Fecha inválida"}))
                ps=_opex_periodos()
                if data.get("borrar"):
                    if len(ps)<=1:
                        return self._send(200,json.dumps({"ok":False,"error":"No se puede borrar la única vigencia"}))
                    ps=[p for p in ps if str(p.get("desde"))!=desde]
                else:
                    if any(str(p.get("desde"))==desde for p in ps):
                        return self._send(200,json.dumps({"ok":False,"error":"Ya existe una vigencia desde esa fecha"}))
                    src=str(data.get("copiar_de") or "")
                    base=[p for p in ps if str(p.get("desde"))==src] or ps[-1:]
                    # arranca como copia de otra vigencia: se cambian solo los rubros que variaron
                    ps.append({"desde":desde,"items":[dict(e) for e in (base[0].get("items") or [])]})
                    ps.sort(key=lambda p:str(p.get("desde")))
                _save("opex.json",ps)
            elif path=="/api/producto":
                pos=str(data.get("pos","")).strip()
                if not pos: return self._send(200,json.dumps({"ok":False,"error":"Falta el nombre"}))
                tipo=data.get("tipo","receta")
                mex=[e for e in _load("maestro_extra.json",[],estricto=True) if str(e.get("pos","")).strip().upper()!=pos.upper()]
                entry={"pos":pos,"cat":data.get("cat") or "GENERICO","canon":data.get("canon") or pos,
                       "tipo":tipo,"factor":float(data.get("factor") or (2 if tipo=="promo_2x1" else 1)),
                       "rend":(float(data["rend"]) if data.get("rend") not in (None,"","0",0) else None),
                       "costeo":"","nota":data.get("nota","") or "Creado desde la app"}
                if tipo in ("receta","promo_2x1"):
                    entry["costeo"]="Receta: "+pos
                    r=_load("recetas_extra.json",{},estricto=True); r[pos]=[[i[0],i[1]] for i in data.get("ingredientes",[])]; _save("recetas_extra.json",r)
                elif tipo in ("botella","pour","directo"):
                    entry["costeo"]="Insumo: "+str(data.get("insumo","")).strip()
                elif tipo=="combo":
                    entry["costeo"]="Combo definido en app"
                    c=_load("combos_extra.json",{},estricto=True); c[pos]=[[x[0],float(x[1]),x[2]] for x in data.get("componentes",[])]; _save("combos_extra.json",c)
                mex.append(entry); _save("maestro_extra.json",mex)
            elif path=="/api/excel":
                dst=generate_excel()
                return self._send(200,json.dumps({"ok":bool(dst),"file":os.path.basename(dst) if dst else ""}))
            elif path=="/api/pull":
                ok,log=run_connector()
                if not ok: return self._send(200,json.dumps({"ok":False,"log":log}))
            else:
                return self._send(404,'{"ok":false}')
        except DatosCorruptos as e:
            return self._send(409,json.dumps({"ok":False,"error":str(e)}))
        except Exception as e:
            # str(e) filtraba rutas absolutas del sistema al cliente. La traza completa va al
            # log local (la ventana de la app), el cliente recibe algo generico.
            print("ERROR en",path,"->",repr(e))
            return self._send(500,json.dumps({"ok":False,"error":"Error interno; mirá la ventana de la app"}))
        ok,log=run_pipeline()
        d=json.load(open(os.path.join(BASE,"datos_dashboard.json"),encoding="utf-8")) if ok else {}
        return self._send(200,json.dumps({"ok":ok,"data":d,"log":log[-400:]}))
    def log_message(self,*a): pass

class Server(socketserver.ThreadingTCPServer):
    allow_reuse_address=True; daemon_threads=True
def main():
    run_pipeline()  # datos frescos al iniciar
    httpd=None; port=None
    for p in range(PORT,PORT+8):
        try: httpd=Server(("127.0.0.1",p),H); port=p; break
        except OSError: continue
    if httpd is None:
        print("No pude abrir ningún puerto (8733-8740). Cerrá otras ventanas de ANTIMO y reintentá.")
        try: input("Enter para cerrar...")
        except Exception: pass
        return
    url=f"http://127.0.0.1:{port}/"
    print("ANTIMO app corriendo en",url,"\n(Dejá esta ventana abierta. Para cerrar la app, cerrá esta ventana.)")
    threading.Timer(1.0,lambda:webbrowser.open(url)).start()
    try: httpd.serve_forever()
    except KeyboardInterrupt: print("\nApp cerrada.")
    finally: httpd.server_close()
if __name__=="__main__": main()
