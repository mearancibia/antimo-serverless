#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ANTIMO - Actualizador v3 INTERACTIVO. Procesa TODOS los rankings de entrada/ como periodos filtrables.
import openpyxl, re, json, os, glob, datetime, zipfile, tempfile, math
BASE=os.path.dirname(os.path.abspath(__file__))
DATOS=os.path.join(BASE,"datos"); ENTRADA=os.path.join(BASE,"entrada")
def _find(patterns, folders):
    for fo in folders:
        for pat in patterns:
            hits=sorted(glob.glob(os.path.join(fo,pat)), key=os.path.getmtime)
            if hits: return hits[-1]
    return None
def _find_all(patterns, folders):
    out=[]
    for fo in folders:
        for pat in patterns:
            for h in glob.glob(os.path.join(fo,pat)):
                if h not in out: out.append(h)
    return sorted(out)
DG=_find(["datos_general.xlsx","datos general.xlsx","*general*.xlsx"],[DATOS,BASE])
RANKS=_find_all(["*anking*.xlsx","*ent*.xlsx","*Vent*.xlsx","api_*.xlsx"],[ENTRADA])
PDF_SRC=_find(["*CAJA*.pdf","*MOVIMIENTO*.pdf","*cierre*.pdf","*.pdf"],[ENTRADA,BASE])
PDF_ALL=_find_all(["*CAJA*.pdf","*MOVIMIENTO*.pdf","*cierre*.pdf","*.pdf"],[ENTRADA])
# Una instalacion nueva todavia no tiene ventas: sin esto salia un traceback de Python, que
# para el dueño de un bar no dice absolutamente nada. Se sale limpio con instrucciones.
if not DG:
    print("\n  Falta el archivo de costos.\n")
    print("  Tiene que estar en:  datos/datos_general.xlsx")
    print("  Si lo moviste o renombraste, volvelo a dejar ahi.\n")
    raise SystemExit(2)
if not RANKS:
    print("\n  Todavia no hay ventas cargadas, asi que no hay nada que calcular.\n")
    print("  Para traerlas:")
    print("    1) Abri ANTIMO (doble clic en run_ANTIMO_app)")
    print("    2) Configura tu cuenta con el boton  Bistrosoft")
    print("    3) Apreta  Traer ventas\n")
    print("  (Tambien podes correr:  python3 conector_bistrosoft.py)\n")
    raise SystemExit(2)
TPL=os.path.join(BASE,"dashboard_tpl.html")
import shutil, contextlib
@contextlib.contextmanager
def ranking_reparado(src):
    """Extrae el xlsx sin mergeCells a un temporal y LO BORRA al salir.
    Antes era mkdtemp() sin limpieza: como el pipeline corre en cada arranque y despues de
    cada POST, se acumulaba un directorio por archivo de entrada/ y por guardado (se llegaron
    a juntar 464, cada uno con un xlsx descomprimido entero)."""
    tmp=tempfile.mkdtemp(prefix="antimo_rank_")
    try:
        dst=os.path.join(tmp,"rank.xlsx")
        with zipfile.ZipFile(src) as z: z.extractall(tmp)
        ws=os.path.join(tmp,"xl","worksheets","sheet1.xml")
        if os.path.exists(ws):
            with open(ws,encoding="utf-8") as f: x=f.read()
            x=re.sub(r"<mergeCells.*?</mergeCells>","",x,flags=re.S); x=re.sub(r"<mergeCell [^>]*/>","",x)
            with open(ws,"w",encoding="utf-8") as f: f.write(x)
        with zipfile.ZipFile(dst,"w",zipfile.ZIP_DEFLATED) as z:
            for root,_,files in os.walk(tmp):
                for f in files:
                    fp=os.path.join(root,f)
                    if fp!=dst: z.write(fp,os.path.relpath(fp,tmp))
        yield dst
    finally:
        shutil.rmtree(tmp,ignore_errors=True)
def _money(s):
    m=re.search(r"-?[\d\.]+,\d{2}", s); return float(m.group(0).replace(".","").replace(",",".")) if m else None
def parse_caja(pdf):
    if not pdf: return None
    try: import pdfplumber
    except Exception:
        # No autoinstalar: --break-system-packages puede alterar el Python del sistema, y una
        # descarga de red no anunciada contradice que la app sea offline. Las cajas ya vienen
        # por la API; el PDF es opcional.
        print("Aviso: los PDF de caja necesitan pdfplumber (pip3 install pdfplumber). Se omiten.")
        return None
    try:
        with pdfplumber.open(pdf) as p: txt="\n".join((pg.extract_text() or "") for pg in p.pages)
    except Exception: return None
    def grab(l):
        m=re.search(re.escape(l)+r".*?(-?[\d\.]+,\d{2})", txt); return _money(m.group(0)) if m else None
    c={"archivo":os.path.basename(pdf)}
    mfe=re.search(r"Fecha de cierre:\s*([\d/: ]+)", txt); c["fecha"]=mfe.group(1).strip() if mfe else ""
    c["total_vendido"]=grab("Total Vendido"); c["efectivo"]=grab("Efectivo:"); c["tarjetas"]=grab("Tarjetas:"); c["qr"]=grab("QR:")
    c["invitaciones"]=grab("Total de Invitaciones"); c["anulaciones"]=grab("Total de anulaciones de Comandas"); c["retiros"]=grab("Total Retiros")
    md=re.search(r"(\d+)\s+DESCUENTOS\s*\$?\s*(-?[\d\.]+,\d{2})", txt); c["descuentos"]=_money(md.group(2)) if md else None
    lines=[l.strip() for l in txt.splitlines()]; ret=[]
    for i,l in enumerate(lines):
        if l.startswith("RETIRO -"):
            mv=_money(l); con=lines[i+1] if i+1<len(lines) else ""
            if mv is not None: ret.append([con[:40],mv])
    c["detalle_retiros"]=ret
    return c
# El workbook se cargaba DOS veces (y Costo_Base se recorria tres): la segunda carga descartaba
# la primera y reconstruia COSTO identico. Una sola pasada.
wb = openpyxl.load_workbook(DG, data_only=True)
COSTO = {}; CB_CAT = {}
for r in list(wb["Costo_Base"].iter_rows(values_only=True))[1:]:
    if r[1] is None: continue
    COSTO[r[1]] = {"precio": r[2], "pres": r[3], "cant_base": r[4], "unidad": r[5], "cxu": r[6]}
    CB_CAT[r[1]] = r[0]

def norm(s):
    return " ".join(str(s).strip().upper().split()) if s is not None else ""
# Lista de Precios (hoja del Excel base): precio OFICIAL, complementario al PVP promedio real
# que sale de las ventas.
# Match en dos pasos, NUNCA fuzzy (Regla #0 — no adivinar a qué producto corresponde un precio):
#   1) por nombre normalizado EXACTO;
#   2) si no hay, por nombre "aplanado": mismas letras y números ignorando espacios, puntuación
#      y tildes. Eso cubre solo diferencias de tipeo del POS ("2 X 1. APEROL" vs "2 x 1 APEROL",
#      "FERNET+COCA" vs "FERNET + COCA", "CUMPLEANOS" vs "CUMPLEAÑOS") — la cadena de letras y
#      dígitos tiene que ser IDÉNTICA, así que NO empareja cosas como "CORONA 330" con
#      "CORONA 33O" (cero vs letra O) ni "COCA 600CC" con "COCA": esos siguen sin match a
#      propósito, porque podrían ser productos distintos y eso lo decide el dueño.
import unicodedata as _ud
def _aplanar(s):
    s=_ud.normalize("NFD",str(s).upper())
    s="".join(c for c in s if _ud.category(c)!="Mn")   # saca tildes/diéresis
    return re.sub(r"[^A-Z0-9]","",s)
PRECIO_LISTA={}; _PL_APL={}; _apl_dup=set()
if "Lista de Precios" in wb.sheetnames:
    for r in list(wb["Lista de Precios"].iter_rows(values_only=True))[1:]:
        if r[1] is None or r[2] is None: continue
        PRECIO_LISTA[norm(r[1])]=r[2]
        a=_aplanar(r[1])
        if a in _PL_APL and _PL_APL[a]!=r[2]: _apl_dup.add(a)   # ambiguo: dos precios distintos
        _PL_APL[a]=r[2]
for a in _apl_dup: _PL_APL.pop(a,None)                          # ante la duda, no matchear
# Equivalencias que NI el match exacto ni el aplanado pueden resolver, porque el nombre cambia
# de verdad: typos del POS ("33O" con letra O en vez de cero, "HEINIKEN") o el Excel detalla el
# envase y el POS no ("COCA 600CC" vs "COCA"). NO se adivinan: cada una fue confirmada por el
# dueño y contrastada contra el precio realmente cobrado (Regla #0). La dif es contra el promedio
# de venta al momento de mapearlas — sirve como evidencia de que son el mismo producto:
#   CORONA 33O      137 u  $6.000 vs $6.000   0,0%
#   COCA            600 u  $3.507 vs $3.500   0,2%
#   SPRITE           55 u  $3.618 vs $3.500   3,4%
#   COCA ZERO        41 u  $3.646 vs $3.500   4,2%
#   FANTA            13 u  $3.731 vs $3.500   6,6%
#   HEINIKEN CHICA   13 u  $4.692 vs $7.000 -33,0%  <- typo claro, pero vendió una sola noche
#                                                      (07-06) con precio promocional. Se mapea
#                                                      igual: esa brecha es justo lo que la
#                                                      columna "Lista" está para mostrar.
PRECIO_LISTA_ALIAS={
    norm("CORONA 33O"):     norm("CORONA 330"),
    norm("COCA"):           norm("COCA 600CC"),
    norm("SPRITE"):         norm("SPRITE 600CC"),
    norm("COCA ZERO"):      norm("COCA ZERO 600CC"),
    norm("FANTA"):          norm("FANTA 600CC"),
    norm("HEINIKEN CHICA"): norm("HEINEKEN CHICA"),
}
def precio_lista_de(k):
    """precio de lista de un producto por su key. None si no hay match seguro."""
    if k in PRECIO_LISTA: return PRECIO_LISTA[k]                # exacto (incluye overrides)
    al=PRECIO_LISTA_ALIAS.get(k)                                # equivalencia confirmada
    if al is not None and al in PRECIO_LISTA: return PRECIO_LISTA[al]
    return _PL_APL.get(_aplanar(k))                             # aplanado, solo si es inequívoco
EQUI = {"cucharada":(12,"g"),"bocha":(60,"g"),"hoja":(0.5,"g"),"gajo":(15,"g"),"rodaja":(12,"g"),
        "medida":(60,"ml"),"trago":(60,"ml"),"shot":(45,"ml"),"lata red bull":(250,"ml"),
        "lata speed":(473,"ml"),"poron":(330,"ml"),"a gusto":(15,"g"),"aceituna":(5,"g")}
ALIAS = {
 "Fernet Branca":"Fernet Branca","Arroz para Sushi":"Arroz base para Sushi","Salmón Fresco":"Proteína para Poke Bowl Salmón","Queso Crema":"Queso Crema con hierbas","Alga Nori":"Alga Nori","Coca Cola (Insumo Barra)":"Coca Cola / Coca Zero",
 "Hielo":"Hielo / Hielo Picado","Hielo picado":"Hielo / Hielo Picado",
 "Gin (Brighton/Beefeater)":"Gin Brighton","Gin":"Gin Brighton","Agua Tónica":"Agua Tónica",
 "Rodaja de Limón":"Limas y Limones (Gajos/Rodajas/Jugos)","Ron Blanco":"Ron Blanco",
 "Soda o Sprite":"Sprite 3L","Sprite":"Sprite 3L","Soda":"Soda fresca","Soda fresca":"Soda fresca",
 "Hojas de Menta":"Menta fresca (Hojas)","Hojas de Menta fresca":"Menta fresca (Hojas)",
 "Lima en gajos":"Limas y Limones (Gajos/Rodajas/Jugos)","Limas en gajos":"Limas y Limones (Gajos/Rodajas/Jugos)",
 "Jugo de Lima":"Limas y Limones (Gajos/Rodajas/Jugos)","Jugo de Limón":"Limas y Limones (Gajos/Rodajas/Jugos)",
 "Azúcar":"Azúcar (Blanca/Almíbar simple)","Azúcar blanca":"Azúcar (Blanca/Almíbar simple)",
 "Almíbar simple":"Azúcar (Blanca/Almíbar simple)","Ron Malibu":"Ron Malibu",
 "Whisky":"Whisky Ballantines","Whisky Red Label":"Whisky Ballantines",
 "Almíbar de Miel":"Almíbar de Miel / Jengibre","Almíbar de Jengibre":"Almíbar de Miel / Jengibre",
 "Cerveza Corona (Porón)":"Cerveza Corona Porroncito 330ml","Tequila":"Tequila","Triple Sec":"Triple Sec",
 "Blue Curaçao":"Blue Curaçao","Vodka Smirnoff":"Vodka Smirnoff","Vodka":"Vodka Smirnoff",
 "Vodka Absolut":"Vodka Absolut","Energizante Speed":"Energizante Speed",
 "Energizante (Speed/Red Bull)":"Energizante Speed","Pulpa de Fruta":"Pulpa / Fruta para Daikiri",
 "Aperol":"Aperol","Espumante (Champagne)":"Espumante / Champagne Chandon",
 "Rodaja de Naranja":"Limas y Limones (Gajos/Rodajas/Jugos)","Cerveza de Jengibre":"Cerveza de Jengibre (Ginger Beer)",
 "Baileys":"Baileys","Helado de crema americana":"Helado (Crema Americana)","Campari":"Campari",
 "Vermouth Rosso":"Vermouth Rosso Cinzano","Vermouth Rosso (Cinzano/Carpano)":"Vermouth Rosso Cinzano",
 "Cachaça":"Cachaça","Cynar":"Cynar","Jugo de Pomelo":"Jugo de Pomelo","Jugo de Naranja o Tónica":"Jugo de Naranja fresco",
 "Lechuga Romana":"Lechuga (Romana/Capuchina)","Lechuga Capuchina picada":"Lechuga (Romana/Capuchina)",
 "Pechuga de Pollo Grillé":"Pechuga de Pollo (Grillé/Crispy)","Pechuga de Pollo Crispy Frito":"Pechuga de Pollo (Grillé/Crispy)",
 "Crotones de Pan":"Crotones de Pan","Queso Parmesano en Hebras":"Queso Parmesano (Hebras/Escamas)",
 "Queso Parmesano rallado":"Queso Parmesano (Hebras/Escamas)","Queso Parmesano en escamas":"Queso Parmesano (Hebras/Escamas)",
 "Aderezo César":"Aderezo César","Arroz base para Sushi":"Arroz base para Sushi",
 "Proteína (Salmón/Pollo)":"Proteína para Poke Bowl Salmón","Palta fresca":"Palta fresca",
 "Pepino japonés":"Pepino (Japonés/Pepinillos en Vinagre)","Pepinillos en vinagre":"Pepino (Japonés/Pepinillos en Vinagre)",
 "Semillas de Sésamo":"Semillas de Sésamo","Salsa de Soja / Teriyaki":"Salsa de Soja / Teriyaki",
 "Pan Baguette Rústico":"Pan Baguette Rústico","Lomito Vacuno o Pechuga":"Lomito Vacuno",
 "Queso Danbo / Muzarella":"Queso Danbo","Tomate en rodajas":"Tomate redondo",
 "Mayonesa clásica":"Mayonesa clásica / Mayonesa base Coleslaw","Mayonesa base Coleslaw":"Mayonesa clásica / Mayonesa base Coleslaw",
 "Mayonesa / Aderezo":"Mayonesa clásica / Mayonesa base Coleslaw","Medallón de Carne Vacuna":"Medallón de Carne Vacuna (Hamburguesas)",
 "Medallón de Carne Vacuna 1":"Medallón de Carne Vacuna (Hamburguesas)","Medallón de Carne Vacuna 2":"Medallón de Carne Vacuna (Hamburguesas)",
 "Pan con Sésamo":"Pan de Hamburguesa (Sésamo/Brioche/Clásico)","Pan Brioche de Hamburguesa":"Pan de Hamburguesa (Sésamo/Brioche/Clásico)",
 "Pan Brioche":"Pan de Hamburguesa (Sésamo/Brioche/Clásico)","Pan de Hamburguesa Clásico":"Pan de Hamburguesa (Sésamo/Brioche/Clásico)",
 "Queso Cheddar en Fetas":"Queso Cheddar (Fetas/Salsa Fundida)","Salsa Cheddar fundido":"Queso Cheddar (Fetas/Salsa Fundida)",
 "Salsa Cheddar Fundido":"Queso Cheddar (Fetas/Salsa Fundida)","Panceta Ahumada crocante":"Panceta Ahumada",
 "Panceta Ahumada picada":"Panceta Ahumada","Panceta picada crocante":"Panceta Ahumada","Panceta Ahumada":"Panceta Ahumada",
 "Huevo frito":"Huevo fresco","Huevo fresco":"Huevo fresco","Cebolla blanca picada":"Cebolla blanca",
 "Cebolla Blanca en juliana":"Cebolla blanca","Ketchup":"Ketchup","Mostaza":"Mostaza",
 "Carne Desmechada (Tapa/Bondiola)":"Carne Desmechada (Tapa de asado/Bondiola)","Pan Ciabatta o Flat":"Pan Ciabatta / Flat / Francés",
 "Pan Francés rústico":"Pan Ciabatta / Flat / Francés","Queso Muzarella fundido":"Queso Muzarella (Barra/Premium)",
 "Queso Muzarella":"Queso Muzarella (Barra/Premium)","Queso Muzarella Premium":"Queso Muzarella (Barra/Premium)",
 "Queso Muzarella en cubos":"Queso Muzarella (Barra/Premium)","Cebolla Caramelizada":"Cebolla Caramelizada",
 "Salsa Especial Park":"Salsa Especial Park","Tapa de Asado Horneada":"Tapa de Asado","Queso Provoleta fundido":"Queso Provoleta",
 "Chimichurri artesanal":"Chimichurri artesanal","Papas Bastón Prefritas":"Papas Fritas (Bastón/Rústicas Prefritas)",
 "Papas Fritas Rústicas":"Papas Fritas (Bastón/Rústicas Prefritas)","Papas Fritas Bastón":"Papas Fritas (Bastón/Rústicas Prefritas)",
 "Ragú de Carne estofada":"Ragú de Carne Estofada","Estofado / Ragú de Carne":"Ragú de Carne Estofada",
 "Aceite de Girasol (absorción)":"Aceite de Girasol (Freír)","Aceite de Girasol (freír)":"Aceite de Girasol (Freír)",
 "Masa de Pizza (bollo)":"Masa de Pizza (Bollos de media masa)","Salsa de Tomate base":"Tomate redondo",
 "Orégano seco":"Orégano seco","Aceitunas Verdes":"Aceitunas (Verdes/Negras)","Aceitunas Verdes y Negras":"Aceitunas (Verdes/Negras)",
 "Aceite de Oliva":"Aceite de Oliva (Común/Extra Virgen)","Aceite de Oliva Extra Virgen":"Aceite de Oliva (Común/Extra Virgen)",
 "Jamón Cocido en fetas":"Jamón Cocido","Jamón Cocido feteado":"Jamón Cocido","Morrones en tiras":"Morrón Rojo",
 "Morrón rojo asado":"Morrón Rojo","Hojas de Albahaca Fresca":"Hojas de Albahaca Fresca","Hojas de Rúcula Fresca":"Hojas de Rúcula Fresca",
 "Jamón Crudo en fetas":"Jamón Crudo","Jamón Crudo feteado":"Jamón Crudo","Cuadrado de Brownie":"Cuadrado de Brownie Chocolate",
 "Bocha de Helado Americana":"Helado (Crema Americana)","Salsa de Chocolate cobertura":"Salsa de Chocolate cobertura",
 "Nueces picadas":"Nueces picadas","Pan Integral / Sin TACC":"Pan Integral / Sin TACC","Zucchini asado":"Zucchini",
 "Berenjena asada":"Berenjena","Hummus (garbanzos)":"Hummus (Garbanzos)","Queso Cream con hierbas":"Queso Crema con hierbas",
 "Sal Fina":"Sal Fina","Sal fina":"Sal Fina","Salsa Barbacoa":"Salsa Barbacoa","Cebolla de Verdeo picada":"Cebolla de Verdeo",
 "Zanahoria rallada":"Zanahoria","Nuggets de Pollo":"Nuggets de Pollo","Repollo blanco y morado":"Repollo (Blanco/Morado)",
 "Maní Salado":"Maní Salado","Queso Pategrás":"Queso Pategrás","Salame tipo Milán":"Salame tipo Milán",
 "Queso Azul":"Queso Azul","Panera (Pancitos variados)":"Panera (Pancitos variados)",
}
FALTANTES = {"Tubos de Calamar","Harina de Trigo"}
def parse_qty(qty):
    if qty is None: return None
    s=str(qty).strip().lower()
    m=re.match(r"^([\d.,]+)\s*(ml|cc)\b",s)
    if m: return (float(m.group(1).replace(",",".")),"ml")
    m=re.match(r"^([\d.,]+)\s*(g|gr|gramos)\b",s)
    if m: return (float(m.group(1).replace(",",".")),"g")
    m=re.match(r"^([\d.,]+)\s*unidad",s)
    if m: return (float(m.group(1).replace(",",".")),"u")
    m=re.match(r"^([\d.,]+)\s*lata",s)
    if m: return ("LATA",float(m.group(1).replace(",",".")))
    for key,(val,ub) in EQUI.items():
        if key in s:
            m2=re.match(r"^([\d.,]+)",s); n=float(m2.group(1).replace(",",".")) if m2 else 1.0
            return (n*val,ub)
    m=re.match(r"^([\d.,]+)$",s)
    if m: return (float(m.group(1).replace(",",".")),"u")
    return None
# gramos/ml que representa "1 unidad" para insumos precificados por g/ml (SUPUESTOS)
UNIDAD_G={"Pan Baguette Rústico":200,"Pan Ciabatta / Flat / Francés":150,
 "Pan de Hamburguesa (Sésamo/Brioche/Clásico)":210,"Pan Integral / Sin TACC":90,
 "Cerveza Corona Porroncito 330ml":330,
 "Limas y Limones (Gajos/Rodajas/Jugos)":12,"Nuggets de Pollo":20}
# g por 1 unidad de aceituna (g-precificado)
UNIT_TO_G={"Aceitunas (Verdes/Negras)":5}
# gramos por 1 pieza, para insumos precificados por unidad ('u') pero recetas en gramos (SUPUESTOS)
PIECE_G={"Medallón de Carne Vacuna (Hamburguesas)":115,"Masa de Pizza (Bollos de media masa)":250,
 "Cuadrado de Brownie Chocolate":120,"Palta fresca":200}
def costo_ingrediente(ing,qty):
    ing=str(ing).strip()
    if ing in FALTANTES: return (None,f"insumo faltante:{ing}")
    insumo=ALIAS.get(ing) or (ing if ing in COSTO else None)
    if insumo is None: return (None,f"sin alias:{ing}")
    if insumo not in COSTO: return (None,f"no en Costo_Base:{insumo}")
    p=parse_qty(qty)
    if p is None: return (None,f"cant no parseada:{qty}")
    cxu=COSTO[insumo]["cxu"]; ub_i=COSTO[insumo]["unidad"]
    # Una celda vacia en la columna cxu de Costo_Base (fila a medio cargar, formula con #DIV/0!)
    # hacia val*None -> TypeError sin capturar, que abortaba TODO el pipeline y dejaba la app sin
    # regenerar. Se trata como cualquier otro dato faltante: N/D para ese producto (Regla #0).
    if cxu is None: return (None,f"insumo sin costo por unidad:{insumo}")
    if p[0]=="LATA":
        if COSTO[insumo]["cant_base"] is None: return (None,f"insumo sin cantidad base:{insumo}")
        return (p[1]*COSTO[insumo]["cant_base"]*cxu,None)
    val,ub=p
    if ub=="u":
        if ub_i=="u": return (val*cxu,None)
        # insumo por g/ml, receta en unidades -> convertir
        if insumo in UNIDAD_G: return (val*UNIDAD_G[insumo]*cxu,None)
        if insumo in UNIT_TO_G: return (val*UNIT_TO_G[insumo]*cxu,None)
        return (None,f"unidad sin peso:{ing}")
    # receta en g/ml
    if ub_i=="u":
        # insumo por unidad, receta en gramos -> usar peso de pieza
        if insumo in PIECE_G: return (val/PIECE_G[insumo]*cxu,None)
        return (None,f"pieza sin peso:{ing}")
    return (val*cxu,None)
RECETAS={}
for sheet,i0,nc in [("Recetas Bebidas",1,0),("Recetas Comida",2,1)]:
    for r in list(wb[sheet].iter_rows(values_only=True))[1:]:
        cells=list(r); nm=cells[nc]
        if not nm: continue
        ings=[]
        for j in range(i0,len(cells)-1,2):
            ing,qty=cells[j],cells[j+1]
            if ing and str(ing).strip(): ings.append((str(ing).strip(),qty))
        RECETAS[norm(nm)]=ings
def costear_receta(rn,factor=1.0):
    ings=RECETAS.get(rn)
    if ings is None: return (None,[f"receta no hallada:{rn}"])
    total=0.0; errs=[]
    for ing,qty in ings:
        c,e=costo_ingrediente(ing,qty)
        if e: errs.append(e)
        elif c: total+=c
    if errs: return (None,errs)
    return (total*factor,[])
MAESTRO={}
for r in list(wb["Maestro_Productos"].iter_rows(values_only=True))[1:]:
    if r[0] is None: continue
    MAESTRO[norm(r[0])]={"cat":r[1],"canon":r[2],"tipo":r[3],"factor":r[4] or 1,"rend":r[5],"costeo":r[6],"nota":r[7]}
MAESTRO[norm("CORONA 33O")]=dict(MAESTRO[norm("Cerveza Corona Porroncito 330ml")])
MAESTRO[norm("CORONA 730")]=dict(MAESTRO[norm("Cerveza Corona Porron 710ml")])
COMBOS={norm("SMIRNOFF + 4 ENERGIZANTES"):[("Vodka Smirnoff",700,"ml"),("Energizante Speed",4,"lata")],
        norm("ABSOLUT + 4 ENERGIZANTES"):[("Vodka Absolut",700,"ml"),("Energizante Speed",4,"lata")],
        norm("BARON B + ENERGIZANTE"):[("Espumante / Champagne Baron B",750,"ml"),("Energizante Speed",2,"lata")],
        norm("SERNOVA + 4 ENERGIZANTES"):[("Vodka Sernova",700,"ml"),("Energizante Speed",4,"lata")]}
def costear_combo(k):
    # componentes editables desde la app: si alguno no existe en Costo_Base (typo, insumo
    # borrado), el combo va a N/D en vez de tirar abajo todo el pipeline.
    t=0.0; errs=[]
    for ins,cant,u in COMBOS[k]:
        if ins not in COSTO: errs.append(f"insumo no encontrado:{ins}"); continue
        if COSTO[ins]["cxu"] is None: errs.append(f"insumo sin costo por unidad:{ins}"); continue
        if u=="ml": t+=cant*COSTO[ins]["cxu"]
        elif u=="lata":
            if COSTO[ins]["cant_base"] is None: errs.append(f"insumo sin cantidad base:{ins}"); continue
            t+=cant*COSTO[ins]["cant_base"]*COSTO[ins]["cxu"]
        else: errs.append(f"unidad no soportada:{u}")
    return (None,errs) if errs else (t,[])
def insumo_pour(c): return c.replace("Insumo:","").strip()
INSUMO_ALIAS={"Gin Beefeater":"Gin Beefeater","Gin Aconcagua":"Gin Aconcagua","Aperol":"Aperol",
 "Whisky Ballantines":"Whisky Ballantines","Jagger":"Jagger","Cerveza Stella 1L":"Cerveza Stella 1L",
 "Cerveza Corona (Porón/Poroncitó)":"Cerveza Corona Porroncito 330ml","Cerveza Corona (proxy)":"Cerveza Corona Porron 710ml",
 "Cerveza Patagonia 730":"Cerveza Patagonia 730","Coca Cola / Coca Zero 600cc":"Coca Cola / Coca Zero 600cc",
 "Sprite 600cc":"Sprite 600cc","Agua Con Gas / Sin Gas":"Agua Con Gas / Sin Gas","Agua Saborizada":"Agua Saborizada",
 "Limonada base":"Limonada base (Clásica/Frutos Rojos/Pomelada)","Energizante Speed":"Energizante Speed",
 "Energizante Red Bull":"Energizante Red Bull","Agua Tónica":"Agua Tónica"}
def costear_producto(k):
    m=MAESTRO.get(k)
    if not m: return {"nd":True,"motivo":"no está en Maestro"}
    t=m["tipo"]
    if t=="sin_datos": return {"nd":True,"motivo":"sin_datos"}
    if t=="combo":
        if k in COMBOS:
            c,errs=costear_combo(k)
            if c is None: return {"nd":True,"motivo":"; ".join(errs[:2])}
            return {"costo":c,"tipo":t}
        return {"nd":True,"motivo":"combo sin composición"}
    if t in ("receta","promo_2x1"):
        rec=str(m["costeo"] or "").replace("Receta:","").strip()
        f=(m.get("factor") or 1)
        c,errs=costear_receta(norm(rec),f)
        if c is None: return {"nd":True,"motivo":"; ".join(errs[:2])}
        return {"costo":c,"tipo":t}
    if t=="botella":
        ins=insumo_pour(str(m["costeo"])); real=INSUMO_ALIAS.get(ins,ins)
        if real in COSTO and COSTO[real]["precio"] is not None: return {"costo":COSTO[real]["precio"],"tipo":t}
        return {"nd":True,"motivo":f"botella sin precio:{ins}" if real in COSTO else f"botella no hallada:{ins}"}
    if t=="pour":
        ins=insumo_pour(str(m["costeo"])); real=INSUMO_ALIAS.get(ins,ins)
        if real in COSTO and m["rend"] and COSTO[real]["cxu"] is not None:
            return {"costo":m["rend"]*COSTO[real]["cxu"],"tipo":t}
        return {"nd":True,"motivo":f"pour sin insumo/rend/costo:{ins}"}
    if t=="directo":
        ins=insumo_pour(str(m["costeo"])); real=INSUMO_ALIAS.get(ins,ins)
        if real in COSTO:
            if m["rend"] and COSTO[real]["cxu"] is not None: return {"costo":m["rend"]*COSTO[real]["cxu"],"tipo":t}
            if not m["rend"] and COSTO[real]["precio"] is not None: return {"costo":COSTO[real]["precio"],"tipo":t}
            return {"nd":True,"motivo":f"directo sin costo:{ins}"}
        return {"nd":True,"motivo":f"directo sin insumo:{ins}"}
    return {"nd":True,"motivo":f"tipo desconocido:{t}"}
# ===== EXPLOSION DE CONSUMO (reposicion) =====
def qty_ingrediente(ing,qty):
    ing=str(ing).strip()
    if ing in FALTANTES: return (None,None,None,f"faltante:{ing}")
    insumo=ALIAS.get(ing) or (ing if ing in COSTO else None)
    if insumo is None or insumo not in COSTO: return (None,None,None,f"sin insumo:{ing}")
    p=parse_qty(qty)
    if p is None: return (None,None,None,f"cant:{qty}")
    ub_i=COSTO[insumo]["unidad"]
    if p[0]=="LATA": return (insumo,p[1]*COSTO[insumo]["cant_base"],ub_i,None)
    val,ub=p
    if ub=="u":
        if ub_i=="u": return (insumo,val,"u",None)
        if insumo in UNIDAD_G: return (insumo,val*UNIDAD_G[insumo],ub_i,None)
        if insumo in UNIT_TO_G: return (insumo,val*UNIT_TO_G[insumo],ub_i,None)
        return (None,None,None,f"unidad sin peso:{ing}")
    if ub_i=="u":
        if insumo in PIECE_G: return (insumo,val/PIECE_G[insumo],"u",None)
        return (None,None,None,f"pieza sin peso:{ing}")
    return (insumo,val,ub_i,None)
def explotar_producto(k):
    m=MAESTRO.get(k)
    if not m: return None
    t=m["tipo"]; out=[]
    if t=="sin_datos": return None
    if t=="combo":
        if k not in COMBOS: return None
        for ins,cant,u in COMBOS[k]:
            if ins not in COSTO: return None  # mismo insumo faltante que en costear_combo -> N/D, no crash
            if u=="ml": out.append((ins,cant,COSTO[ins]["unidad"]))
            elif u=="lata": out.append((ins,cant*COSTO[ins]["cant_base"],COSTO[ins]["unidad"]))
        return out
    if t in ("receta","promo_2x1"):
        rec=norm(str(m["costeo"] or "").replace("Receta:","").strip())
        ings=RECETAS.get(rec)
        if ings is None: return None
        f=(m.get("factor") or 1)
        for ing,qty in ings:
            insumo,qb,ub,err=qty_ingrediente(ing,qty)
            if err: return None
            out.append((insumo,qb*f,ub))
        return out
    if t=="botella":
        ins=insumo_pour(str(m["costeo"])); real=INSUMO_ALIAS.get(ins,ins)
        if real not in COSTO: return None
        return [(real,COSTO[real]["cant_base"],COSTO[real]["unidad"])]
    if t=="pour":
        ins=insumo_pour(str(m["costeo"])); real=INSUMO_ALIAS.get(ins,ins)
        if real not in COSTO or not m["rend"]: return None
        return [(real,m["rend"],COSTO[real]["unidad"])]
    if t=="directo":
        ins=insumo_pour(str(m["costeo"])); real=INSUMO_ALIAS.get(ins,ins)
        if real not in COSTO: return None
        q=m["rend"] if m["rend"] else COSTO[real]["cant_base"]
        return [(real,q,COSTO[real]["unidad"])]
    return None

# ============ MODELO UNIFICADO (linea de tiempo dia a dia) ============
import re as _re
UNIFICAR={norm("speed"):norm("SPEED"),norm("redbull"):norm("REDBULL"),
 norm("Gin aconcagua vaso"):norm("Gin Aconcagua"),
 norm("VASO LIMONADA FRUTOS ROJOS"):norm("LIMONADA FRUTOS ROJOS"),
 norm("VASO LIMONADA FR"):norm("LIMONADA FRUTOS ROJOS"),
 norm("GIN BRIGTON"):norm("Gin Brighton")}  # typo del POS, confirmado por el dueño: mismo producto
FOOD_CATS={"HAMBURGUESAS Y SANDWICH","PARA PICAR Y PAPAS","PIZZAS","ENSALADAS","SIN TACC Y VEGGIE","POSTRES"}
def grupo_cat(cat): return "COMIDA" if str(cat).upper() in FOOD_CATS else "BEBIDA"
DOW=["Lun","Mar","Mie","Jue","Vie","Sab","Dom"]

def nd_guia(nombre, motivo):
    up=_re.sub(r"[^A-Z0-9 ]","",str(nombre).upper())
    toks=[x for x in up.split() if len(x)>3]
    if "combo" in motivo:
        return ("Falta definir la composición (qué botellas/unidades incluye).","Confirmámelo y lo cargo en Maestro_Productos.")
    if motivo.startswith("no está en Maestro"):
        cost_hit=None
        for name in COSTO:
            words=set(_re.sub(r"[^A-Z0-9 ]"," ",name.upper()).split())
            if toks and any(t in words for t in toks): cost_hit=name; break
        rec_hit=any(toks and (set(rn.split()) & set(toks)) for rn in RECETAS)
        if rec_hit or cost_hit:
            base="su "+("receta ya existe" if rec_hit else ("costo ya existe: "+cost_hit))
            return ("No está en Maestro_Productos ("+base+").","Agregá una fila en Maestro_Productos de datos_general.xlsx.")
        return ("No está en Maestro_Productos y no encuentro su costo/receta.","Cargá el insumo en 'Costo items' y agregá la fila en Maestro_Productos.")
    if "faltante" in motivo or "sin_datos" in motivo or "receta no" in motivo:
        return ("Falta la receta y/o el costo de algún insumo.","Cargá los insumos que faltan en 'Costo items'.")
    return (motivo,"Revisar en datos_general.xlsx.")

# ---- fusiones (extienden MAESTRO/RECETAS/COSTO antes de costear) ----
_ext=os.path.join(DATOS,"maestro_extra.json")
if os.path.exists(_ext):
    for e in json.load(open(_ext,encoding="utf-8")):
        MAESTRO[norm(e["pos"])]={"cat":e["cat"],"canon":e["canon"],"tipo":e["tipo"],
            "factor":e.get("factor") or 1,"rend":e.get("rend"),"costeo":e.get("costeo"),"nota":e.get("nota","")}
    print("Maestro extendido con",len(json.load(open(_ext,encoding="utf-8"))),"productos")
# rendimiento (ml/porcion) de pours editado desde la app: override parcial sobre MAESTRO,
# no reemplaza la fila entera (a diferencia de maestro_extra.json), asi que funciona igual
# para productos que vienen del Excel y para los creados desde la app.
_pex=os.path.join(DATOS,"pours_extra.json")
if os.path.exists(_pex):
    for k,rend in json.load(open(_pex,encoding="utf-8")).items():
        if k in MAESTRO: MAESTRO[k]=dict(MAESTRO[k]); MAESTRO[k]["rend"]=rend
_rex=os.path.join(DATOS,"recetas_extra.json")
if os.path.exists(_rex):
    for nm,ings in json.load(open(_rex,encoding="utf-8")).items():
        RECETAS[norm(nm)]=[(i[0],i[1]) for i in ings]
_iex=os.path.join(DATOS,"insumos_extra.json")
if os.path.exists(_iex):
    for e in json.load(open(_iex,encoding="utf-8")):
        nm=e["nombre"]; base=dict(COSTO.get(nm,{}))
        cant=e.get("cant_base", base.get("cant_base")) or 1
        precio=e.get("precio", base.get("precio"))
        cxu=e.get("cxu")
        if cxu is None and precio is not None and cant: cxu=precio/cant
        COSTO[nm]={"precio":precio,"pres":e.get("pres",base.get("pres","")),
                   "cant_base":cant,"unidad":e.get("unidad",base.get("unidad","")),"cxu":cxu}
        CB_CAT[nm]=e.get("cb_cat",CB_CAT.get(nm,"Comidas - Sushi"))
# override de PRECIOS de insumos existentes (editor de costos)
_pov=os.path.join(DATOS,"precios_override.json")
if os.path.exists(_pov):
    for nm,precio in json.load(open(_pov,encoding="utf-8")).items():
        if nm in COSTO and COSTO[nm].get("cant_base"):
            COSTO[nm]=dict(COSTO[nm]); COSTO[nm]["precio"]=precio
            COSTO[nm]["cxu"]=precio/COSTO[nm]["cant_base"]
# combos editados (editor de promos)
_cbx=os.path.join(DATOS,"combos_extra.json")
if os.path.exists(_cbx):
    for kpos,comp in json.load(open(_cbx,encoding="utf-8")).items():
        COMBOS[norm(kpos)]=[(c[0],c[1],c[2]) for c in comp]
# precio de lista editado desde la app: pisa el valor de la hoja "Lista de Precios" del Excel
# sin tocarla. Vacio/0 desde el endpoint significa "volver al de la hoja" (se maneja en app_antimo).
_plo=os.path.join(DATOS,"precio_lista_override.json")
if os.path.exists(_plo):
    try:
        for k,precio in json.load(open(_plo,encoding="utf-8")).items(): PRECIO_LISTA[k]=precio
    except Exception: pass
# precios sospechosos: marca del dueño sobre productos mal cargados en el POS.
# NO altera ningun costo ni margen: solo viaja al tablero como etiqueta.
SOSP={}
_ssp=os.path.join(DATOS,"sospechosos.json")
if os.path.exists(_ssp):
    try: SOSP=json.load(open(_ssp,encoding="utf-8"))
    except Exception: SOSP={}
# noches que el bar NO abrio: las marca el dueño para que no figuren como dato faltante.
CERRADOS={}
_dce=os.path.join(DATOS,"dias_cerrados.json")
if os.path.exists(_dce):
    try: CERRADOS=json.load(open(_dce,encoding="utf-8"))
    except Exception: CERRADOS={}
# stock: conteos manuales del dueño (solo insumos que el marca). El front proyecta el consumo
# real (consumo_dia, mas abajo) desde la fecha del conteo para estimar cuanto queda.
STOCK={}
_stk=os.path.join(DATOS,"stock.json")
if os.path.exists(_stk):
    try: STOCK=json.load(open(_stk,encoding="utf-8"))
    except Exception: STOCK={}

def detalle_producto(k):
    """desglose de costo por 1 unidad: lista de insumos con cantidad, costo unit y subtotal."""
    exp=explotar_producto(k)
    if not exp: return []
    out=[]
    for insumo,qb,ub in exp:
        cxu=COSTO.get(insumo,{}).get("cxu",0) or 0
        out.append({"insumo":insumo,"qty":round(qb,2),"unidad":ub,"cxu":round(cxu,3),"sub":round(qb*cxu,1)})
    return out

# ---- recolectar todas las ventas en una linea de tiempo ----
def _ym_de(path):
    m=_re.search(r"(20\d\d)-(\d\d)", os.path.basename(path))
    return (m.group(1),m.group(2)) if m else None
prods={}; dias={}; consumo_dia={}; split={}
for path in RANKS:
    ym=_ym_de(path)
    try:
        # las filas se materializan DENTRO del with: al salir se borra el temporal
        with ranking_reparado(path) as _rp:
            rows=list(openpyxl.load_workbook(_rp,data_only=True).active.iter_rows(values_only=True))
    except Exception as e: print("WARN",os.path.basename(path),e); continue
    hdr=None; header=[]
    for i,row in enumerate(rows[:6]):
        vals=[str(c).strip().lower() if c is not None else "" for c in row]
        if "nombre" in vals and ("vendidos" in vals or "monto" in vals): hdr=i; header=vals; break
    if hdr is None: continue
    def col(*n):
        for x in n:
            if x in header: return header.index(x)
        return None
    cN,cV,cM,cF=col("nombre"),col("vendidos"),col("monto"),col("fecha")
    for r in rows[hdr+1:]:
        if not r or cN is None or r[cN] is None: continue
        nombre=r[cN]
        try: u=int(r[cV] or 0)
        except (TypeError,ValueError): u=0
        try: monto=float(r[cM] or 0)
        except (TypeError,ValueError): monto=0.0
        if u==0 and monto==0: continue
        # El ranking trae la fecha como "DD-MM" (sin año); el año sale del nombre del archivo
        # (api_ventas_YYYY-MM.xlsx). Si no se puede deducir, se descarta el año en vez de inventar
        # uno: antes había un "2026" hardcodeado que iba a quedar mal el año que viene.
        fecha=str(r[cF]) if cF is not None and r[cF] is not None else "s/f"
        dd,mm=(fecha.split("-")+["",""])[:2]
        yy=ym[0] if ym else ""
        iso=f"{yy}-{mm}-{dd}" if (yy and mm and dd) else ""
        # CLAVE del día = fecha ISO completa, no "DD-MM". Los archivos de ventas se acumulan en
        # entrada/ (uno por mes, para siempre), así que en 2027 van a convivir api_ventas_2026-07
        # y api_ventas_2027-07: con la clave vieja, el "18-07" de cada año se sumaba en un solo
        # día sin ningún aviso. `fecha` se conserva solo como etiqueta para mostrar.
        dkey=iso or fecha
        if dkey not in dias:
            try: dow=DOW[datetime.date.fromisoformat(iso).weekday()]
            except Exception: dow=""
            dias[dkey]={"fecha":fecha,"iso":iso,"dow":dow}
        k=norm(nombre); k=UNIFICAR.get(k,k)
        p=prods.setdefault(k,{"raw":nombre,"byday":{}})
        bd=p["byday"].setdefault(dkey,[0,0.0]); bd[0]+=u; bd[1]+=monto
        exp=explotar_producto(k)
        if exp:
            g=grupo_cat(MAESTRO.get(k,{}).get("cat",""))
            cd=consumo_dia.setdefault(dkey,{})
            for insumo,qb,ub in exp:
                cd[insumo]=cd.get(insumo,0.0)+qb*u
                s=split.setdefault(insumo,{"BEBIDA":0.0,"COMIDA":0.0}); s[g]+=qb*u

# ---- costear cada producto ----
productos=[]
for k,p in prods.items():
    cc=costear_producto(k)
    cat=MAESTRO.get(k,{}).get("cat","(sin cat)"); grupo=grupo_cat(cat)
    byday={f:[bd[0],round(bd[1])] for f,bd in p["byday"].items()}
    # "key" = nombre normalizado y unificado: es la identidad estable del producto
    # (el POS puede escribirlo distinto). Con ella se guardan las marcas del dueño.
    base={"pos":p["raw"],"key":k,"cat":cat,"grupo":grupo,"byday":byday,"precio_lista":precio_lista_de(k)}
    _sp=SOSP.get(k) or {}
    if _sp.get("estado"):
        base["susp"]=_sp["estado"]; base["susp_motivo"]=_sp.get("motivo","") or ""
    if cc.get("nd"):
        fa,do=nd_guia(p["raw"],cc["motivo"])
        base.update({"nd":True,"motivo":cc["motivo"],"falta":fa,"donde":do})
    else:
        base.update({"nd":False,"tipo":cc["tipo"],"costo":round(cc["costo"],1),
                     "nota":MAESTRO.get(k,{}).get("nota","") or "","breakdown":detalle_producto(k)})
        _m=MAESTRO.get(k,{}); _t=cc["tipo"]
        if _t in ("receta","promo_2x1"):
            _rn=str(_m.get("costeo") or "").replace("Receta:","").strip()
            base["receta_nombre"]=_rn
            base["receta_ings"]=[[i[0],i[1]] for i in RECETAS.get(norm(_rn),[])]
        elif _t=="combo":
            _cc=COMBOS.get(k)
            if _cc: base["combo_comp"]=[[x[0],x[1],x[2]] for x in _cc]
        base["editable"]=_t in ("receta","promo_2x1","combo")
    productos.append(base)

# ---- metadata de insumos (para compras) ----
insumos_meta={}
for insumo in set(i for cd in consumo_dia.values() for i in cd):
    c=COSTO.get(insumo,{}); s=split.get(insumo,{"BEBIDA":0,"COMIDA":0})
    insumos_meta[insumo]={"cxu":c.get("cxu",0),"precio":c.get("precio",0),"cant_base":c.get("cant_base",0),"present":c.get("pres",""),
        "unidad":c.get("unidad",""),"cb_cat":CB_CAT.get(insumo,"Otros"),
        "grupo":"COMIDA" if s["COMIDA"]>s["BEBIDA"] else "BEBIDA","compartido":s["BEBIDA"]>0 and s["COMIDA"]>0}
consumo_dia={f:{i:round(q,2) for i,q in cd.items()} for f,cd in consumo_dia.items()}
dias_list=sorted(dias.values(), key=lambda d:d["iso"] or d["fecha"])

# ---- cajas (PDF + API) ----
cajas=[]
for _pdf in PDF_ALL:
    ck=parse_caja(_pdf)
    if ck:
        mfk=_re.search(r"(\d{2})/(\d{2})/(\d{4})", ck.get("fecha",""))
        if mfk: ck["fecha_key"]=mfk.group(0); ck["fecha_dia"]=mfk.group(1)+"/"+mfk.group(2)
        else: ck["fecha_key"]=""
        cajas.append(ck)
_capi=os.path.join(DATOS,"cajas_api.json")
if os.path.exists(_capi):
    try:
        for c in json.load(open(_capi,encoding="utf-8")): cajas.append(c)
    except Exception: pass
# índice DD-MM -> ISO, solo para las noches en que ese DD-MM es INEQUÍVOCO. Si el histórico
# llega a tener el mismo día-mes en dos años, esa entrada se descarta y no se adivina.
_ddmm2iso={}; _ddmm_amb=set()
for _d in dias.values():
    _f=_d.get("fecha"); _i=_d.get("iso")
    if not _f or not _i: continue
    if _f in _ddmm2iso and _ddmm2iso[_f]!=_i: _ddmm_amb.add(_f)
    _ddmm2iso[_f]=_i
for _f in _ddmm_amb: _ddmm2iso.pop(_f,None)
def _caja_iso(c):
    """Fecha de la noche en ISO. El conector nuevo la guarda completa en 'fecha_iso'; los
    registros viejos y el PDF traen 'DD-MM' o 'DD/MM/YYYY' y hay que deducirla."""
    fi=str(c.get("fecha_iso") or "")
    if _re.match(r"^\d{4}-\d{2}-\d{2}$", fi): return fi          # dato explícito: no se adivina
    m=_re.search(r"(\d{2})/(\d{2})/(\d{4})", str(c.get("fecha_key") or "") or str(c.get("fecha") or ""))
    if m: return "%s-%s-%s"%(m.group(3),m.group(2),m.group(1))
    f=str(c.get("fecha") or "")
    if f in _ddmm2iso: return _ddmm2iso[f]                        # DD-MM que solo existe en un año
    return ""
# Un mismo cierre puede llegar por PDF y por API a la vez: se queda el de la API,
# que trae comensales, descuentos y detalle por operador.
_uni={}; _sin=[]
for c in cajas:
    c["iso"]=_caja_iso(c)
    if not c["iso"]: _sin.append(c); continue
    prev=_uni.get(c["iso"])
    if prev is None or (c.get("archivo")=="Bistrosoft API" and prev.get("archivo")!="Bistrosoft API"):
        _uni[c["iso"]]=c
_dup=len(cajas)-len(_uni)-len(_sin)
cajas=sorted(_uni.values(), key=lambda c:c["iso"])+_sin
if _dup: print("Cajas duplicadas descartadas (PDF ya cubierto por la API):",_dup)

# ---- OPEX (desde opex.json editable; se siembra del Excel la 1ra vez) ----
# Fecha de la vigencia inicial: bien atras, para que cubra todo el historico que exista.
OPEX_DESDE_0="2000-01-01"
opex_total=0; opex_pend=0; opex_detalle=[]; opex_periodos=[]
_cero_ok=set()
_ce=os.path.join(DATOS,"opex_cero_confirmado.json")
if os.path.exists(_ce): _cero_ok={str(x).strip().lower() for x in json.load(open(_ce,encoding="utf-8"))}
_opx=os.path.join(DATOS,"opex.json")
if not os.path.exists(_opx):
    seed=[]
    try:
        _oov={}
        _ov=os.path.join(DATOS,"opex_override.json")
        if os.path.exists(_ov): _oov={str(k).strip().lower():v for k,v in json.load(open(_ov,encoding="utf-8")).items()}
        for r in wb["Costos FIJOS (OPEX)"].iter_rows(min_row=2,values_only=True):
            if not r or r[0] is None: continue
            if str(r[0]).upper().startswith("TOTAL"): continue
            if r[1] is None: continue
            cat=r[0]; item=r[1]; cant=r[2]; unit=r[3]; monto=r[5] or 0
            key=str(item).strip().lower()
            if key in _oov: monto=_oov[key]
            if cant and unit and abs((cant or 0)*(unit or 0)-monto)<1 and monto: cantidad=cant; unitario=unit
            else: cantidad=1; unitario=monto
            seed.append({"cat":cat,"item":item,"cantidad":cantidad,"unitario":unitario,"confirmado_cero":key in _cero_ok})
        json.dump([{"desde":OPEX_DESDE_0,"items":seed}],open(_opx,"w",encoding="utf-8"),ensure_ascii=False,indent=1)
    except Exception: pass
def _opex_calc(items):
    """Devuelve (detalle, total, pendientes) de una lista de rubros."""
    det=[]; tot=0; pend=0
    for e in (items or []):
        cantidad=e.get("cantidad",1) or 0; unitario=e.get("unitario",0) or 0; monto=round(cantidad*unitario)
        key=str(e.get("item","")).strip().lower()
        cz=bool(e.get("confirmado_cero")) or (key in _cero_ok)
        det.append({"cat":e.get("cat","(sin cat)"),"item":e.get("item",""),"cantidad":cantidad,
                    "unitario":unitario,"monto":monto,"confirmado_cero":cz})
        if monto==0:
            if not cz: pend+=1
        else: tot+=monto
    return det,tot,pend
# opex.json admite dos formatos: lista plana (viejo, una sola foto) o lista de vigencias
# [{desde, items}]. La lista plana se lee como una vigencia que rige desde siempre, asi que
# los tableros viejos siguen dando exactamente el mismo numero.
_raw=[]
try: _raw=json.load(open(_opx,encoding="utf-8")) or []
except Exception: _raw=[]
if _raw and isinstance(_raw[0],dict) and "items" in _raw[0]: _pers=[p for p in _raw if isinstance(p,dict)]
elif _raw: _pers=[{"desde":OPEX_DESDE_0,"items":_raw}]
else: _pers=[]
_pers.sort(key=lambda p:str(p.get("desde") or OPEX_DESDE_0))
opex_periodos=[]
for p in _pers:
    det,tot,pend=_opex_calc(p.get("items"))
    opex_periodos.append({"desde":str(p.get("desde") or OPEX_DESDE_0),"opex":tot,"opex_pend":pend,"opex_detalle":det})
# DATA.opex / opex_detalle / opex_pend = la vigencia que rige HOY (compatibilidad)
_hoy=datetime.date.today().isoformat(); _vig=None
for p in opex_periodos:
    if p["desde"]<=_hoy: _vig=p
if _vig is None and opex_periodos: _vig=opex_periodos[0]
if _vig: opex_total=_vig["opex"]; opex_pend=_vig["opex_pend"]; opex_detalle=_vig["opex_detalle"]
if not opex_detalle: opex_total=10460000

# ---- logo (si existe logo.png/.jpg en la carpeta) ----
logo_uri=""
import base64 as _b64, mimetypes as _mt
for _n in ["logo.png","logo.jpg","logo.jpeg","logo.svg","logo.webp"]:
    for _fo in [BASE,DATOS]:
        _lp=os.path.join(_fo,_n)
        if os.path.exists(_lp):
            _m=_mt.guess_type(_lp)[0] or "image/png"
            logo_uri="data:%s;base64,%s"%(_m,_b64.b64encode(open(_lp,"rb").read()).decode())
            break
    if logo_uri: break
DATA={"generado":datetime.date.today().isoformat(),"logo":logo_uri,"opex":opex_total,"opex_pend":opex_pend,
      "opex_detalle":opex_detalle,"dias":dias_list,"productos":productos,
      "insumos":insumos_meta,"consumo_dia":consumo_dia,"cajas":cajas,"dias_cerrados":CERRADOS,"stock":STOCK,
      "opex_periodos":opex_periodos}
with open(TPL,encoding="utf-8") as f: tpl=f.read()
# encoding explicito: sin el, open() usa el del locale y un "Cachaça"/"CUMPLEAÑOS" revienta a
# mitad de escritura, dejando truncado justo el archivo que la app sirve.
with open(os.path.join(BASE,"dashboard_ANTIMO.html"),"w",encoding="utf-8") as f:
    f.write(tpl.replace("@@DATA@@", json.dumps(DATA,ensure_ascii=False)))
with open(os.path.join(BASE,"datos_dashboard.json"),"w",encoding="utf-8") as f:
    json.dump(DATA,f,ensure_ascii=False)
print(f"OK. Dias:{len(dias_list)} Productos:{len(productos)} (N/D {sum(1 for p in productos if p.get('nd'))}) Insumos:{len(insumos_meta)} Cajas:{len(cajas)} OPEX:${opex_total:,.0f}")
